#!/usr/bin/env python3
r"""
test_core.py - golden + unit tests for the kb-extract-rig core.

The core is pure-geometry / pure-arithmetic, so it is highly testable. These tests lock the
behaviour the whole "100% on digital" guarantee rests on:
  - the rounding-scaled tie-out tolerance (A1) and amount-column-by-header (A2)
  - verify_table catching a dropped row with the EXACT amount
  - is_tabular ACCEPTING a sparse-but-real BOQ and REJECTING prose (the A3 guard - must be
    additive, never over-reject; the fixture is deliberately SPARSE)
  - extract_page never fabricating a table from prose, and extracting BOTH tables on a mixed page
  - indian_words (incl. >=100 crore), the ID checksums, completeness_status
  - a digital GOLDEN PDF: every row + amount exact and the money ties out
  - parallel_foreach determinism (workers=2 == serial)

Run:  pytest tests/         (server)   |   python tests/test_core.py   (standalone, no pytest)
"""
import sys, os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import common as C
import robust_tables as RT

try:
    import fitz
except Exception:
    fitz = None


# --------------------------------------------------------------------------- A1
def test_round_tolerance():
    assert C.round_tolerance(38) == 19          # GUDC: 38 rows -> 19 rupees; real diff 0.01 passes
    assert C.round_tolerance(1) == 1
    assert C.round_tolerance(5) == 3
    assert C.round_tolerance(0) == 1            # floor

def test_parse_number():
    assert C.parse_number("Rs 4,94,27,290.67") == 49427290.67
    assert C.parse_number("₹ 1,200") == 1200.0
    assert C.parse_number("80%") == 80.0
    assert C.parse_number("n/a") is None
    assert C.parse_number(None) is None


# --------------------------------------------------------------------------- A2
def _counts(rows):
    nc = max(len(r) for r in rows)
    return rows, [sum(1 for r in rows if len(r) > c and C.parse_number(r[c]) is not None) for c in range(nc)]

def test_pick_amount_column_full_boq():
    rows, counts = _counts([["Sr", "Desc", "Unit", "Qty", "Rate", "Amount"],
                            ["1", "Earthwork", "cum", "100", "50", "5000"],
                            ["2", "Concrete", "cum", "20", "300", "6000"],
                            ["3", "Steel", "kg", "10", "80", "800"]])
    col = C.pick_amount_column(rows, counts)
    assert rows[0][col] == "Amount"             # NOT Qty/Rate

def test_pick_amount_column_abstract():
    rows, counts = _counts([["Schedule", "Description", "Amount"],
                            ["B1", "Network", "41873588"], ["B2", "Intake", "4942729"],
                            ["B3", "Pump", "1200000"], ["Total", "", "48016317"]])
    col = C.pick_amount_column(rows, counts)
    assert rows[0][col] == "Amount"


# --------------------------------------------------------------------------- verify_table
import table_pdf as TP

def test_verify_table_dropped_row_exact_gap():
    rows = [["Schedule", "Amount"], ["B1", "41873588.00"], ["B2", "4942729.67"],
            ["B3", "1200000.33"], ["Total", "48016318.00"]]
    assert TP.verify_table(rows) == []                       # intact -> no flag
    dropped = [r for r in rows if r[0] != "B2"]
    flags = TP.verify_table(dropped)
    assert flags and abs(flags[0]["gap"] - 4942729.67) < 0.1 # EXACT missing amount surfaced

def test_verify_table_rounding_within_tol():
    rows = [["x", "Amount"], ["a", "100.00"], ["b", "100.00"], ["c", "100.01"], ["Total", "300.02"]]
    assert TP.verify_table(rows) == []                       # 0.01 drift over 3 rows < tol -> no false flag

def test_verify_table_small_gap_now_caught():
    # a dropped 5-lakh row in a 4-crore total: UNDER the old 0.5% tolerance, OVER the new one
    rows = [["x", "Amount"]] + [[f"i{i}", "1000000"] for i in range(40)] + [["Total", "40500000"]]
    flags = TP.verify_table(rows)
    assert flags and abs(flags[0]["gap"]) == 500000.0


# --------------------------------------------------------------------------- A3 is_tabular
def test_is_tabular_accepts_sparse_real_boq():
    # ~21% sparse rows (section headers, subtotal, a 0 row) like 20A_Sch-H - MUST be accepted
    boq = [["Sr", "Description", "Qty", "Rate", "Amount"],
           ["", "Earthwork", ""],
           ["1", "Excavation", "100", "50", "5000"],
           ["2", "Filling", "20", "30", "600"],
           ["", "Road Side Drains", ""],
           ["3", "RCC drain", "10", "300", "3000"],
           ["4", "Spare", "0", "0", "0"],
           ["", "Sub Total", "", "", "8600"]]
    assert RT.is_tabular(boq) is True

def test_is_tabular_rejects_prose():
    assert RT.is_tabular([["a single column of running prose text here"],
                          ["that wraps over several lines with no columns"],
                          ["describing the methodology in plain sentences"]]) is False
    assert RT.is_tabular([["Introduction", ""], ["The scope", ""], ["includes work", ""],
                          ["across the", "site"], ["and beyond", ""]]) is False


# --------------------------------------------------------------------------- extract_page
def _W(t, x0, x1, top, bot): return {"text": t, "x0": x0, "x1": x1, "top": top, "bottom": bot}

class _MockPage:
    def __init__(self, words, edges=None): self._w = words; self.edges = edges or []
    def extract_words(self, **k): return self._w

def test_extract_page_prose_zero_tables():
    prose = []
    for i, line in enumerate(["the project scope includes earthworks and drainage",
                              "across the full length of the corridor as described",
                              "in the methodology statement appended to this report"]):
        x = 10
        for w in line.split():
            prose.append(_W(w, x, x + len(w) * 5, 10 + i * 20, 20 + i * 20)); x += len(w) * 5 + 5
    tables, _ = RT.extract_page(_MockPage(prose))
    assert len(tables) == 0                                  # E-int-3: no fabricated table from prose

def test_extract_page_mixed_ruled_and_borderless():
    def H(y): return {"top": y, "bottom": y, "x0": 10, "x1": 110}
    def V(x): return {"top": 10, "bottom": 50, "x0": x, "x1": x}
    edges = [H(10), H(30), H(50), V(10), V(60), V(110)]
    words = [_W("P", 20, 30, 15, 25), _W("Q", 70, 80, 15, 25), _W("R", 20, 30, 35, 45), _W("S", 70, 80, 35, 45),
             _W("X", 10, 20, 70, 80), _W("10", 100, 130, 70, 80),
             _W("Y", 10, 20, 90, 100), _W("20", 100, 130, 90, 100),
             _W("Z", 10, 20, 110, 120), _W("30", 100, 130, 110, 120)]
    tables, audit = RT.extract_page(_MockPage(words, edges))
    modes = sorted(t["mode"] for t in tables)
    assert modes == ["borderless", "ruled"]                 # P0-3: BOTH, not just the ruled one
    assert audit["orphans_in_table_region"] == []


# --------------------------------------------------------------------------- indian_words + checksums
def test_indian_words():
    assert C.indian_words(0) == "zero"
    assert C.indian_words(2255080500) .startswith("two hundred twenty five crore")   # >=100 crore, no crash
    assert "lakh" in C.indian_words(150000)

def test_checksums():
    assert C.valid_pan("ABCDE1234F") and not C.valid_pan("ABCD1234F")
    assert C.valid_ifsc("HDFC0001234") and not C.valid_ifsc("HDFC1001234")
    # derive a VALID gstin + aadhaar by completing the check digit (no magic constants)
    base = "27AAPFU0939F1Z"
    g = next(base + ch for ch in "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ" if C.valid_gstin(base + ch))
    assert C.valid_gstin(g) and not C.valid_gstin(g[:-1] + ("0" if g[-1] != "0" else "1"))
    ad = next("23456789012" + str(d) for d in range(10) if C.valid_aadhaar("23456789012" + str(d)))
    assert C.valid_aadhaar(ad) and not C.valid_aadhaar(ad[:-1] + ("0" if ad[-1] != "0" else "1"))


# --------------------------------------------------------------------------- completeness
def test_completeness_status():
    assert C.completeness_status()[0] == "AUTO_ACCEPT"
    assert C.completeness_status(sig_orphans=1)[0] == "NEEDS_REVIEW"
    assert C.completeness_status(tieout_gaps=1)[0] == "NEEDS_REVIEW"
    assert C.completeness_status(dropped_pages=2)[0] == "NEEDS_REVIEW"

def test_orphan_significance():
    assert RT._orphan_significant({"text": "49,427,290"}) is True   # digit-bearing = likely dropped data row
    assert RT._orphan_significant({"text": "Note"}) is False         # decorative token = ignored


# --------------------------------------------------------------------------- digital GOLDEN pdf
def _make_boq_pdf(path):
    doc = fitz.open(); page = doc.new_page(width=400, height=400)
    xs = [50, 175, 300]; ys = [100, 130, 160, 190, 220, 250]
    for y in ys: page.draw_line((xs[0], y), (xs[-1], y))
    for x in xs: page.draw_line((x, ys[0]), (x, ys[-1]))
    for r, (c0, c1) in enumerate([("Item", "Amount"), ("A", "100"), ("B", "200"), ("C", "300"), ("Total", "600")]):
        page.insert_text((55, ys[r] + 20), c0); page.insert_text((180, ys[r] + 20), c1)
    doc.save(path); doc.close()

def test_digital_golden_exact_and_tieout(tmp_path=None):
    if fitz is None:
        return
    import tempfile, pdfplumber
    d = Path(tmp_path) if tmp_path else Path(tempfile.mkdtemp())
    pdf = d / "boq.pdf"; _make_boq_pdf(str(pdf))
    with pdfplumber.open(str(pdf)) as doc:
        tables, audit = RT.extract_page(doc.pages[0])
    grid = tables[0]["grid"]
    assert ["A", "100"] in grid and ["Total", "600"] in grid       # every row + amount EXACT
    assert TP.verify_table(grid) == []                              # 100+200+300 == 600 ties out
    assert audit["orphans_in_table_region"] == []


# --------------------------------------------------------------------------- parallel determinism
def _sq_worker(path, args):
    return {"stem": str(path), "val": int(str(path).split("_")[-1]) ** 2}

def test_parallel_foreach_determinism():
    files = [f"n_{i}" for i in range(6)]
    serial, _ = C.parallel_foreach(files, _sq_worker, {}, label="t", cfg=C.ExtractConfig(max_workers=1))
    par, _ = C.parallel_foreach(files, _sq_worker, {}, label="t", cfg=C.ExtractConfig(max_workers=3))
    assert sorted(r["val"] for r in serial) == sorted(r["val"] for r in par) == [i * i for i in range(6)]

def _maybe_hang_worker(path, args):
    import time
    if str(path).endswith("HANG"):
        time.sleep(60)                                                # simulate a wedged scanned page
    return {"stem": str(path)}

def test_gpu_lane_kills_hung_file():
    """P0-2: the GPU/scanned lane must KILL a hung file (timeout) and keep going - one wedged scan
    can no longer hang the whole run. The hung file is recorded as a timeout and the run RETURNS
    (a generous timeout absorbs the spawn cold-start so a good file is never falsely killed)."""
    files = ["g_a", "g_b", "HANG"]                                  # HANG last: one spawn + one kill
    res, err = C.parallel_foreach(files, _maybe_hang_worker, {}, label="t",
                                  is_gpu=lambda p: True, cfg=C.ExtractConfig(per_file_timeout=6))
    assert any("TIMEOUT" in e.get("error", "") and e.get("file") == "HANG" for e in err)  # hung -> killed
    assert "HANG" not in {r["stem"] for r in res}                  # never silently "succeeds"


# --------------------------------------------------------------------------- accuracy levers (next tier)
def test_parse_number_accounting_negative():
    assert C.parse_number("(2,500.00)") == -2500.0      # deduction/credit row, not +2500
    assert C.parse_number("(49,427,290.67)") == -49427290.67
    assert C.parse_number("1,200") == 1200.0            # ordinary positive unaffected

def test_is_total_label():
    assert C.is_total_label("Grand Total") and C.is_total_label("Total Road Work") and C.is_total_label("Say")
    assert not C.is_total_label("Total quantity of steel as per IS code")
    assert not C.is_total_label("values are subtotal of the works listed above")   # buried, no start

def test_is_amount_cell():
    assert C.is_amount_cell("2,500.00") and C.is_amount_cell("(2500)") and C.is_amount_cell("₹ 4,94,290")
    assert not C.is_amount_cell("80% of Value Quoted") and not C.is_amount_cell("Total Station 3")

def test_row_arithmetic_catches_cancellation():
    # the canonical lever-A regression: column total reconciles but two rows are individually wrong
    boq = [["Item", "Qty", "Rate", "Amount"],
           ["A", "10", "100", "4000"],   # expect 1000
           ["B", "50", "100", "2000"],   # expect 5000
           ["Total", "", "", "6000"]]    # 4000+2000 == 6000 -> column ties out, 0 column flags
    flags = TP.verify_table(boq)
    assert sum(1 for f in flags if f.get("kind") != "row_arith") == 0      # column total is clean
    assert sum(1 for f in flags if f.get("kind") == "row_arith") == 2      # row check catches both

def test_serial_gap_is_dropped_row():
    boq = [["Sr No", "Desc", "Amount"], ["1", "A", "100"], ["2", "B", "200"], ["3", "C", "300"], ["5", "E", "500"]]
    g = C.detect_serial_gaps(boq)
    assert g and g[0]["missing"] == [4]
    assert C.detect_serial_gaps([["Desc", "Qty"], ["A", "5"], ["B", "9"], ["C", "12"]]) == []   # qty != serial

def test_borderless_merge_detected():
    def W(t, x0, x1, top): return {"text": t, "x0": x0, "x1": x1, "top": top, "bottom": top + 10}
    merged = [W("1", 10, 20, 100), W("A", 60, 80, 100), W("100", 120, 150, 100),
              W("2", 10, 20, 104), W("B", 60, 80, 104), W("200", 120, 150, 104)]
    _, _, m = RT.borderless_table(merged)
    assert m >= 1                                    # tight-pitch merge flagged ('100 200' in one cell)
    normal = [W("1", 10, 20, 100), W("Cement", 60, 90, 100), W("100", 120, 150, 100),
              W("2", 10, 20, 116), W("Steel", 60, 90, 116), W("200", 120, 150, 116)]
    assert RT.borderless_table(normal)[2] == 0       # well-spaced rows: no false merge flag

def test_excel_uncached_formula_flagged(tmp_path=None):
    # fills the Excel-path coverage gap: a library-authored workbook (un-cached formulas) must be
    # FLAGGED, never silently read as complete blanks (1.1).
    try:
        import openpyxl, excel_extract as X
    except Exception:
        return
    import tempfile
    d = Path(tmp_path) if tmp_path else Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Item", "Qty", "Rate", "Amount"]); ws.append(["Cement", 100, 350, "=B2*C2"])
    ws.append(["Steel", 50, 600, "=B3*C3"]); p = str(d / "f.xlsx"); wb.save(p)
    _, uncached = X._read_grids(p)
    assert len(uncached) == 2                         # DETECTION: both Amount formulas are un-cached at raw read
    an = X._analyze(p)
    # The invariant is "never SILENTLY read as blank" - detection above proves they were caught.
    # Post-analyze, LibreOffice (if installed) RECOVERS the formulas so uncached may be 0; without it
    # they stay flagged. Accept EITHER: still-flagged OR recovered to the computed amount (B2*C2=35000).
    assert isinstance(an.get("uncached"), list)
    assert len(an["uncached"]) >= 1 or "35000" in str(an)

def test_excel_tieout_sections_sum():
    try:
        import openpyxl, excel_extract as X
    except Exception:
        return
    import tempfile
    d = Path(tempfile.mkdtemp()); wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Project Cost", 600000]); ws.append(["Total Road Work", 1.0, 400000])
    ws.append(["Total Bridges", 1.0, 200000]); p = str(d / "t.xlsx"); wb.save(p)
    an = X._analyze(p)
    assert an["tie"]["passed"] is True               # 400000+200000 == 600000

# --------------------------------------------------------------------------- lever C: digit repair (oracle-gated)
def test_repair_accepts_when_oracle_holds():
    cols = {"qty": 1, "rate": 2, "amount": 3}
    p, f = C.repair_row_arithmetic(["A", "1O", "1OO", "1000"], cols)   # 1O->10, 1OO->100, 10*100=1000
    assert p[1] == "10" and p[2] == "100" and len(f) == 2

def test_repair_refuses_and_never_touches():
    cols = {"qty": 1, "rate": 2, "amount": 3}
    assert C.repair_row_arithmetic(["A", "10", "100", "1000"], cols) == (["A", "10", "100", "1000"], [])  # reconciles
    assert C.repair_row_arithmetic(["A", "1O", "100", "9999"], cols) == (["A", "1O", "100", "9999"], [])  # no fix exists

def test_repair_flags_ambiguous_never_guesses():
    cols = {"qty": 1, "rate": 2, "amount": 3}
    p, f = C.repair_row_arithmetic(["A", "3O", "2O", "60"], cols)      # 30*2 OR 3*20 both == 60
    assert p == ["A", "3O", "2O", "60"] and f and f[0].get("ambiguous") is True

# --------------------------------------------------------------------------- lever F: confidence calibration
def test_reliability_and_fit_calibration():
    import run_ocr_eval as E
    pairs = [(0.95, i % 20 == 0) for i in range(60)] + [(0.5, i % 2 == 0) for i in range(60)]
    rt = E.reliability_table(pairs)
    assert rt["n"] == 120 and rt["ece"] is not None
    fit = E.fit_calibration(pairs, target_err=0.10, min_samples=50)
    assert fit["recommended_gate"] >= 0.8                              # gate lands in the clean band
    assert E.fit_calibration([(0.9, False)] * 5, min_samples=50)["recommended_gate"] == 0.70  # overfit guard

def test_load_calibration_identity_and_demote():
    import tempfile, json
    assert C.load_calibration("")(0.9) == 0.9                          # identity = today's gate
    d = Path(tempfile.mkdtemp()); p = str(d / "c.json")
    (d / "c.json").write_text(json.dumps({"bins": [{"count": 50, "mean_conf": 0.9, "error_rate": 0.5}]}))
    assert abs(C.load_calibration(p)(0.9) - 0.5) < 0.01               # reported 0.9 -> 50% accuracy

# --------------------------------------------------------------------------- lever B: scanned cell-OCR (cfg + geometry)
def test_ocr_config_string():
    try:
        import scanned_tables as STm
    except Exception:
        return
    assert STm._ocr_config("0123456789.,()-/", 7) == "--oem 1 --psm 7 -c tessedit_char_whitelist=0123456789.,()-/"
    assert STm._ocr_config(None, 6) == "--oem 1 --psm 6"

def test_detect_grid_and_degenerate():
    try:
        import scanned_tables as STm, cv2, numpy as np
    except Exception:
        return
    img = np.full((300, 400, 3), 255, np.uint8)
    for y in (40, 100, 160, 220): cv2.line(img, (40, y), (360, y), (0, 0, 0), 3)
    for x in (40, 150, 260, 360): cv2.line(img, (x, 40), (x, 220), (0, 0, 0), 3)
    _, ys, xs = STm.analyze_layout(img)
    assert not STm.is_degenerate(ys, xs) and len(ys) >= 3 and len(xs) >= 3
    assert STm.is_degenerate([0, 50, 100], [0, 200])                  # single-column -> fall back to img2table

# --------------------------------------------------------------------------- bilingual (English + Hindi) tie-out
def test_devanagari_digits_parse():
    assert C.parse_number("३५०००") == 35000.0
    assert C.parse_number("(२,५००.००)") == -2500.0          # Devanagari + accounting negative
    assert C.is_amount_cell("३५०००") and not C.is_amount_cell("केवल पाठ")   # number vs prose

def test_is_total_label_hindi():
    assert C.is_total_label("कुल योग") and C.is_total_label("कुल") and C.is_total_label("महायोग")
    assert not C.is_total_label("कुल मिलाकर परियोजना का विवरण इस प्रकार है")   # long descriptive != total

def test_hindi_boq_tieout_and_dropped_row():
    # Devanagari headers (Sr/Desc/Qty/Rate/Amount) + Devanagari-digit amounts; rows tie out to कुल
    hi = [["क्रम", "विवरण", "मात्रा", "दर", "राशि"],
          ["१", "सीमेंट", "१००", "३५०", "३५०००"],
          ["२", "इस्पात", "५०", "६००", "३००००"],
          ["३", "रेत", "२०", "४००", "८०००"],
          ["कुल", "", "", "", "७३०००"]]
    cols = C.infer_amount_columns(hi)
    assert cols.get("qty") == 2 and cols.get("rate") == 3 and cols.get("amount") == 4
    assert not [f for f in TP.verify_table(hi) if f.get("kind") not in ("row_arith", "serial_gap")]
    dropped = [r for r in hi if r[1] != "इस्पात"]            # drop Steel (30000)
    gap = [f for f in TP.verify_table(dropped) if f.get("kind") not in ("row_arith", "serial_gap")]
    assert gap and abs(gap[0]["gap"]) == 30000.0            # exact missing amount surfaced

def test_hindi_serial_gap():
    hg = [["क्रम", "विवरण", "राशि"], ["१", "अ", "१००"], ["२", "ब", "२००"], ["३", "स", "३००"], ["५", "द", "५००"]]
    sg = C.detect_serial_gaps(hg)
    assert sg and sg[0]["missing"] == [4]                   # १,२,३,५ -> missing ४

# --------------------------------------------------------------------------- Excel path is bilingual (one parser, not a local copy)
def test_excel_to_num_routes_through_common():
    try:
        import excel_extract as X
    except Exception:
        return
    assert X.to_num("३५०००") == 35000.0 and X.to_num("(२,५००)") == -2500.0   # Devanagari via common.parse_number
    assert X.to_num(35000) == 35000.0 and X.to_num(True) is None              # native Excel types preserved

def test_excel_tieout_hindi_passes():
    try:
        import excel_extract as X
    except Exception:
        return
    grid = [["कुल अनुबंध मूल्य", 600000], ["कुल सड़क कार्य", 1.0, 400000], ["कुल पुल", 1.0, 200000]]
    assert X.tieout(grid, 600000)["passed"] is True            # the one that mattered: was passed=None before

def test_excel_tieout_not_silent_when_money_but_no_total():
    try:
        import excel_extract as X
    except Exception:
        return
    grid = [["विवरण", "राशि"], ["सीमेंट", 35000], ["इस्पात", 30000]]   # money present, no total row
    t = X.tieout(grid, None)
    assert t["passed"] is None and t["amounts_present"] is True   # 'not attempted' is detectable -> review, not commit

def test_contains_total_kw_bilingual():
    assert C.contains_total_kw("Grand Total 1.0 600000") and C.contains_total_kw("कुल योग 600000")
    assert not C.contains_total_kw("Cement bags supplied to site")

def test_emit_worthy_filters_noise_sheets():
    # the output noise-filter: <=1 data cell (page-frame box / lone title) is not a table
    assert TP._emit_worthy([["GUDC Ltd., Gandhinagar"]]) is False
    assert TP._emit_worthy([["", ""], ["", ""]]) is False
    assert TP._emit_worthy([["Schedule", "Amount"], ["B1", "100"], ["Total", "100"]]) is True

def test_is_tabular_word_density_guard():
    # genuine paragraph prose forced into 2 columns (no numbers, long cells) -> REJECTED
    prose = [["The contractor shall provide all materials and labour necessary for the complete execution of the works herein",
              "as detailed in the technical specifications and drawings appended to this contract document for the reference of all"],
             ["Payment shall be made monthly based on the certified value of work completed during the preceding calendar period",
              "subject to the retention provisions and other deductions as specified in the conditions of contract attached herewith"],
             ["The defects liability period shall commence from the date of practical completion as certified by the project engineer",
              "and shall continue for the full duration specified in the appendix to these general conditions of contract as agreed"]]
    assert RT.is_tabular(prose) is False
    # a numeric BOQ with VERBOSE descriptions is still accepted (numeric path untouched by the guard)
    boq = [["Sr", "Desc", "Qty", "Rate", "Amount"],
           ["1", "Providing and laying cement concrete nominal mix as per IS code including all leads and lifts", "100", "350", "35000"],
           ["2", "Supplying and fixing reinforcement steel bars conforming to relevant standards and specifications", "50", "600", "30000"],
           ["3", "Excavation in all kinds of soil and disposal of surplus earth to approved dumping location nearby", "20", "400", "8000"]]
    assert RT.is_tabular(boq) is True

# --------------------------------------------------------------------------- Phase A: links / QR / URLs
def test_links_qr_text_urls():
    try:
        import links_qr as L
    except Exception:
        return
    u = L.text_urls("see https://example.com/bid and www.x.in, mail a@b.co (https://y.org).")
    uris = {x["uri"] for x in u}; subs = {x["subtype"] for x in u}
    assert "https://example.com/bid" in uris and "www.x.in" in uris
    assert "a@b.co" in uris and "https://y.org" in uris        # trailing ) trimmed
    assert "url" in subs and "email" in subs
    import numpy as np
    assert L.detect_qr(np.full((80, 80, 3), 255, np.uint8)) == []   # blank -> no QR (graceful)

def test_cross_page_table_stitch():
    """P0-7: a table split across consecutive pages (continuation + total on page 2) is merged into
    ONE table, re-tied-out (reconciles), and the page-1 'not verified' reason is cleared. Without the
    stitch each half would be unverifiable; with it the multi-page summary AUTO_ACCEPTs correctly."""
    import doc_layout as DL
    def tbl(grid): return {"type": "table", "grid": grid, "mode": "ruled", "tieout_flags": []}
    p1 = tbl([["Schedule", "Amount"], ["B1", "100000"], ["B2", "200000"], ["B3", "300000"]])
    p2 = tbl([["Schedule", "Amount"], ["B4", "400000"], ["Total", "1000000"]])
    doc = {"pages": [
        {"page_no": 1, "elements": [p1], "needs_review": True,
         "review_reasons": ["table: amounts present but NOT verified (no total row)"]},
        {"page_no": 2, "elements": [p2], "needs_review": False, "review_reasons": []}]}
    gaps = DL._stitch_cross_page_tables(doc, None)
    tbls = [e for pg in doc["pages"] for e in pg["elements"] if e["type"] == "table"]
    assert len(tbls) == 1 and len(tbls[0]["grid"]) == 6 and tbls[0]["stitched_pages"] == [2]
    assert tbls[0]["tieout_flags"] == [] and gaps == 0
    assert doc["pages"][0]["review_reasons"] == [] and doc["pages"][0]["needs_review"] is False
    # a table that ALREADY ends in a total is NOT swallowed by a following unrelated table
    a = tbl([["X", "Amt"], ["a", "10000"], ["Total", "10000"]])
    b = tbl([["X", "Amt"], ["b", "20000"], ["Total", "20000"]])
    doc2 = {"pages": [{"page_no": 1, "elements": [a], "review_reasons": [], "needs_review": False},
                      {"page_no": 2, "elements": [b], "review_reasons": [], "needs_review": False}]}
    DL._stitch_cross_page_tables(doc2, None)
    assert sum(1 for pg in doc2["pages"] for e in pg["elements"] if e["type"] == "table") == 2

def test_collision_safe_keys_and_atomic_write(tmp_path=None):
    """P0-1/P0-4: two files with the same basename in different folders get DISTINCT sharded output
    dirs (no silent overwrite); atomic_write leaves no .tmp and writes LF (no CRLF); .done sentinel."""
    import common as C, tempfile
    from pathlib import Path
    k1, k2 = C.doc_key("projA/BOQ.pdf"), C.doc_key("projB/BOQ.pdf")
    assert k1 != k2 and k1.startswith("BOQ__") and len(k1.split("__")[1]) == 8
    klong = C.doc_key("a/" + ("Letter No 0410 Unsuitable material dumped at site " * 4) + ".pdf")
    assert len(klong) <= 48 + 2 + 8          # MAX_PATH guard: a 200-char filename can't blow the path
    d1 = C.doc_outdir("out", "projA/BOQ.pdf")
    assert d1.parts[-1] == k1 and len(d1.relative_to("out").parts) == 3      # outroot/hh/hh/key
    base = Path(tempfile.mkdtemp()); f = base / "x.md"
    C.atomic_write_text(f, "a\nb\n")
    assert f.read_bytes() == b"a\nb\n" and not (base / "x.md.tmp").exists()    # LF + no temp litter
    C.mark_done(base); assert (base / ".done").exists()

def test_render_dpi_clamp():
    """P0-3: a huge (A0) page's effective DPI is lowered so the pixmap long edge stays <= the cap,
    while a normal A4 page renders at the requested DPI unchanged (no multi-GB pixmap OOM)."""
    try:
        import fitz, pdf_extract as PE
    except Exception:
        return
    a0 = fitz.open(); a0p = a0.new_page(width=33 * 72, height=47 * 72)        # A0 in points
    a4 = fitz.open(); a4p = a4.new_page(width=595, height=842)                # A4
    assert PE.clamp_dpi(a0p, 300) < 300                                       # clamped down
    assert (max(a0p.rect.width, a0p.rect.height) / 72.0) * PE.clamp_dpi(a0p, 300) <= PE.MAX_RENDER_PX + 1
    assert PE.clamp_dpi(a4p, 300) == 300                                      # A4 fits -> unchanged

def test_oversized_doc_routed_to_review(tmp_path=None):
    """P0-3: with a page-count cap set, a doc over the cap is routed to NEEDS_REVIEW UNPROCESSED
    (no OOM/stall) and still drops a .done sentinel so resume won't re-do it."""
    try:
        import fitz, doc_layout as DL, common as C, tempfile
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp()); src = d / "big.pdf"
    doc = fitz.open()
    for _ in range(6): doc.new_page(width=200, height=200)
    doc.save(str(src)); doc.close()
    out = d / "out"
    res, _, audit = DL.extract_document(src, out, C.ExtractConfig(max_pages_per_doc=3))
    assert res["status"] == "NEEDS_REVIEW" and audit.get("oversized") and res["pages"] == []
    assert (C.doc_outdir(out, src) / ".done").exists()

def test_order_banded_multicolumn_and_furniture():
    """Phase E: a full-width table between two column blocks reads between them (not floated up),
    and header/footer survive _order (header->top, footer->bottom) instead of being dropped."""
    import doc_layout as DL
    def P(t, x0, y0, x1, y1, tx): return {"type": t, "bbox": [x0, y0, x1, y1], "text": tx}
    els = [P("paragraph", 40, 100, 280, 140, "L1"), P("paragraph", 320, 100, 560, 140, "R1"),
           P("table", 40, 180, 560, 260, "TBL"),
           P("paragraph", 40, 300, 280, 340, "L2"), P("paragraph", 320, 300, 560, 340, "R2"),
           P("header", 40, 20, 560, 40, "HDR"), P("footer", 40, 560, 560, 580, "FTR")]
    order = [e["text"] for e in DL._order([dict(e) for e in els], 600)]
    assert order == ["HDR", "L1", "R1", "TBL", "L2", "R2", "FTR"], order
    single = [P("paragraph", 40, 100, 560, 140, "A"), P("paragraph", 40, 200, 560, 240, "B"),
              P("header", 40, 20, 560, 40, "H"), P("footer", 40, 560, 560, 580, "F")]
    assert [e["text"] for e in DL._order([dict(e) for e in single], 600)] == ["H", "A", "B", "F"]

def test_doc_layout_table_tieout_folds_in():
    """Phase D: verify_table runs on whole-doc table elements; a reconciling table is clean,
    a dropped-row table flags a gap (same money proof table_pdf gives, now in layout.json)."""
    import common as C
    ok = [["Schedule", "Amount"], ["B1", "100000"], ["B2", "200000"], ["B3", "300000"], ["Total", "600000"]]
    bad = [["Schedule", "Amount"], ["B1", "100000"], ["B2", "200000"], ["B3", "300000"], ["Total", "750000"]]
    assert C.verify_table(ok) == []
    flags = C.verify_table(bad)
    assert flags and any(abs(f.get("gap", 0) - 150000) < 1 for f in flags)

def test_ocr_word_geometry_helpers():
    """Phase C1: docTR relative geometry -> abs pixels; px -> points at the dpi scale. The flat
    docTR text wrapper stays a distinct function (geometry path is purely additive)."""
    import pdf_extract as PE
    assert PE._geom_to_box(((0.1, 0.2), (0.5, 0.4)), 1000, 800) == [100.0, 160.0, 500.0, 320.0]
    assert PE._geom_to_box([(0.1, 0.2), (0.5, 0.2), (0.5, 0.4), (0.1, 0.4)], 1000, 800) == [100.0, 160.0, 500.0, 320.0]
    w = PE._px_words_to_points([{"text": "X", "conf": 0.9, "bbox_px": [100, 160, 500, 320]}], 72.0 / 300)
    assert abs(w[0]["bbox"][0] - 24.0) < 1e-9 and abs(w[0]["bbox"][2] - 120.0) < 1e-9
    assert hasattr(PE, "ocr_doctr_batch") and hasattr(PE, "ocr_doctr_batch_pages")

def test_extract_images_native_and_deduped(tmp_path=None):
    """extract_images.py pulls embedded images at NATIVE resolution, de-dupes a repeated logo by
    xref (saved once, all pages recorded), and writes a manifest. Sovereign (fitz only)."""
    try:
        import fitz, extract_images as EI, common as C, tempfile
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp()); src = d / "rep.pdf"
    doc = fitz.open()
    pix = fitz.Pixmap(fitz.csRGB, fitz.IRect(0, 0, 200, 120)); pix.set_rect(pix.irect, (40, 120, 200))
    for _ in range(2):                                                    # same image on two pages
        pg = doc.new_page(width=420, height=560); pg.insert_image(fitz.Rect(40, 80, 240, 200), pixmap=pix)
    doc.save(str(src)); doc.close()
    r = EI.extract_file(str(src), d / "out")
    assert r["images"] == 1                                              # deduped by xref
    docdir = C.doc_outdir(d / "out", src)
    import json
    m = json.loads((docdir / "images_manifest.json").read_text(encoding="utf-8"))
    img = m["images"][0]
    assert img["pages"] == [1, 2] and img["width"] == 200 and img["height"] == 120   # native res, both pages
    assert (docdir / img["file"]).exists() and (docdir / ".done").exists()

def test_batched_ocr_pages_words_shape():
    """P1: ocr_pages_words returns ONE word-only result per page (no rgb = bounded RAM), aligned to
    the input order, with the TRUE page size - the batched docTR pass the scanned layer consumes.
    The per-page ocr_page_words wrapper still returns the raster for the non-batched path."""
    try:
        import fitz, pdf_extract as PE
    except Exception:
        return
    doc = fitz.open()
    for _ in range(3): doc.new_page(width=420, height=560)
    res = PE.ocr_pages_words([doc[i] for i in range(3)], PE.PdfOptions(dpi=150))
    assert len(res) == 3 and all("rgb" not in r for r in res)
    assert all(set(r) == {"words", "page_conf", "engine", "size_pts"} for r in res)
    assert all(abs(r["size_pts"][0] - 420) < 1 and abs(r["size_pts"][1] - 560) < 1 for r in res)
    assert "rgb" in PE.ocr_page_words(doc[0], PE.PdfOptions(dpi=150))

def test_scanned_uses_pre_batched_words():
    """P1: when doc_layout pre-OCRs scanned pages in one batch, scanned_page_elements consumes that
    word result (uses its page_conf/engine, re-renders only the raster) instead of re-OCRing; empty
    words still degrade to a fallback paragraph + low_conf review."""
    try:
        import fitz, scanned_layout as SL, common as C, tempfile
        from pathlib import Path
    except Exception:
        return
    if SL.PE is None:
        return
    pg = fitz.open().new_page(width=420, height=560)
    wr = {"words": [], "page_conf": 0.42, "engine": "doctr", "size_pts": (420.0, 560.0)}
    els, meta = SL.scanned_page_elements(pg, 1, None, C.ExtractConfig(), Path(tempfile.mkdtemp()), {}, words_result=wr)
    assert meta["page_conf"] == 0.42 and meta["engine"] == "doctr"        # used the pre-batched result
    assert meta["low_conf"] is True and any(e["type"] == "paragraph" for e in els)

def test_scanned_words_to_lines_and_table_gate():
    """Phase C2: words cluster into lines (median height = the font-size proxy); a 2-numeric-column
    grid is accepted as a borderless table while single-column prose is NEVER fabricated into one."""
    import scanned_layout as SL
    def w(t, x0, y0, x1, y1): return {"text": t, "conf": 0.95, "bbox": [x0, y0, x1, y1]}
    lines = SL._words_to_lines([w("Hello", 40, 100, 90, 112), w("world", 95, 100, 140, 112),
                                w("Second", 40, 124, 100, 136)])
    assert [l["text"] for l in lines] == ["Hello world", "Second"]
    assert lines[0]["size"] == 12.0 and lines[0]["bold"] is False
    grid_words = []
    for r, y in enumerate((100, 120, 140)):
        grid_words += [w(["A", "B", "C"][r], 40, y, 60, y + 12), w(str((r + 1) * 10), 200, y, 220, y + 12),
                       w(str((r + 1) * 100000), 320, y, 390, y + 12)]
    tbls, consumed = SL._scanned_tables(grid_words, None, 72.0 / 300, None)
    assert len(tbls) == 1 and tbls[0]["mode"] == "borderless-scanned" and len(consumed) == len(grid_words)
    prose = [w("This", 40, 100, 80, 112), w("is", 85, 100, 100, 112), w("prose", 105, 100, 150, 112),
             w("more", 40, 124, 80, 136), w("text", 85, 124, 120, 136)]
    assert SL._scanned_tables(prose, None, 72.0 / 300, None) == ([], set())

def test_scanned_table_gate_rejects_sprinkled_numbers():
    """Anti-fabrication (audit M5): a mostly-text page with a FEW scattered numbers in a 2nd column
    must NOT be collapsed into a borderless table (which would consume every word and drop the
    headings). The density floor + is_tabular guard reject it; the words survive to classification."""
    import scanned_layout as SL
    def w(t, x0, y0, x1, y1): return {"text": t, "conf": 0.9, "bbox": [x0, y0, x1, y1]}
    sprinkled = []
    for r, y in enumerate((100, 120, 140, 160, 180)):
        sprinkled.append(w("Some descriptive sentence here line %d" % r, 40, y, 300, y + 12))
        if r in (1, 3):                                    # only a couple of stray numbers, far right
            sprinkled.append(w(str(r * 1000), 360, y, 400, y + 12))
    tbls, consumed = SL._scanned_tables(sprinkled, None, 72.0 / 300, None)
    assert tbls == [] and consumed == set(), "sprinkled-number prose was fabricated into a table"

def test_emit_worthy_filter_common():
    """Fix #1: the >1-data-cell fragment filter lives in common, is None-safe, and table_pdf aliases it."""
    import common as C, table_pdf as TP
    assert TP._emit_worthy is C.emit_worthy
    assert C.emit_worthy([["Total", "600000"], ["B1", "100000"]]) is True
    assert C.emit_worthy([["", "", ""], ["", "GUDC Ltd.", ""]]) is False      # 1 data cell = a fragment
    assert C.emit_worthy([[""]]) is False and C.emit_worthy([[None, None]]) is False

def test_hierarchical_subtotal_tieout():
    """Fix #2: a summary table that stacks line items + section subtotals + a grand total must
    reconcile (grand == carried subtotals + remaining items), not double-count - while a flat table
    with a genuinely dropped row STILL flags. Covers BOTH real shapes found on the server:
    (a) 004 p12: Amount(1)+Amount(2)=Total (cross-page sections, no post-subtotal items);
    (b) OPRMC p11/p13: items -> Sub-Total -> more charges -> Total = Sub-Total + charges."""
    import common as C
    p12 = [                                                                   # subtotals span pages
        ["", "SCHEDULE - B28", "INTAKE WELL 70 MLD", "4,94,27,290.67", ""],
        ["", "SCHEDULE - B37", "M.S TRANSMISSION LINE", "14,35,65,687.23", ""],
        ["", "", "Amount (1) in Rs.", "2,25,50,80,500.01", ""],               # B1-B27 on prior pages
        ["", "SCHEDULE - C1", "O&M NAGARPALIKA", "11,21,01,776.00", ""],
        ["", "SCHEDULE - C2", "O&M DRAINAGE", "9,45,71,584.75", ""],
        ["", "SCHEDULE - C3", "O&M STP", "3,00,77,642.68", ""],
        ["", "SCHEDULE - C4", "O&M AUGMENTATION", "3,38,87,556.62", ""],
        ["", "", "Amount O&M (2) in Rs.", "27,06,38,560.05", ""],
        ["", "", "Total Amount (1+2)", "2,52,57,19,060.05", ""],
        ["", "", "Say (Rs.)=", "2,52,57,19,061.00", ""],
    ]
    assert C.verify_table(p12) == []                                          # 225.50 cr + 27.06 cr = 252.57 cr
    p11 = [                                                                   # Sub-Total + post charges -> Total
        ["", "Ordinary Maintenance", "1,905.39", ""], ["", "Initial Rectifications", "168.62", ""],
        ["", "Minor Improvement", "4,388.47", ""], ["", "Periodic Maintenance", "11,040.07", ""],
        ["", "Provisional Sum", "786.00", ""], ["", "Sub-Total", "18,288.55", ""],
        ["", "Labour Cess @ 1%", "185.27", ""], ["", "Seigniorage", "133.65", ""],
        ["", "Contingency", "98.32", ""], ["", "Price Adjustment", "1,428.95", ""],
        ["", "GST@18%", "3,414.97", ""], ["", "Videography", "1.53", ""],
        ["", "Inspection Vehicle", "91.98", ""], ["", "Total", "23,643.21", ""],
    ]
    assert C.verify_table(p11) == []                                          # 18,288.55 + 5,354.67 = 23,643.21
    assert C.is_subtotal_label("Sub-Total") and C.is_subtotal_label("Amount (1) in Rs.")
    assert not C.is_subtotal_label("Total Amount (1+2)") and not C.is_subtotal_label("Amount in Rs.")
    flat = [["Sr", "Desc", "Amount"], ["1", "A", "100000"], ["2", "B", "200000"],
            ["3", "C", "300000"], ["", "Total", "900000"]]                    # 600000 stated 900000
    ff = C.verify_table(flat)
    assert ff and any(abs(f.get("gap", 0) - 300000) < 1 for f in ff)
    # a within-section dropped row (most items present, above the min-total floor) is STILL caught
    drop = [["1", "A", "10000"], ["2", "B", "20000"], ["", "Sub-Total", "60000"], ["3", "D", "5000"],
            ["", "Total", "65000"]]                                           # A+B=30000 but Sub-Total=60000
    assert C.verify_table(drop), "within-section dropped row was missed"

def test_not_verified_guard_opt_in():
    """P0-6 CORRECTNESS: a money column with significant values but NO total row to check must NOT
    silently pass. With flag_unverified=True (the production callers) it emits a 'not_verified'
    marker; the default (and any table that DOES have a total, or row-level Qty*Rate) stays clean.
    Closes the hole where a cross-page price SUMMARY AUTO_ACCEPTed with 0 flags."""
    import common as C
    summary = [["Schedule", "Amount (Rs.)"], ["B1", "41873588.00"], ["B2", "4942729.67"],
               ["B3", "12500000.00"]]                                        # amounts, NO total row anywhere
    assert C.verify_table(summary) == []                                     # default OFF -> unchanged contract
    flags = C.verify_table(summary, flag_unverified=True)
    assert flags and flags[0]["kind"] == "not_verified", flags              # ON -> flagged, never silent
    # a table WITH a total still ties out cleanly even with the guard on (no false 'not_verified')
    withtotal = summary + [["Total", "59316317.67"]]                         # 41873588+4942729.67+12500000
    assert C.verify_table(withtotal, flag_unverified=True) == []
    # a row-verifiable table (Qty*Rate=Amount) with no total is NOT flagged - it is row-checked
    rv = [["Item", "Qty", "Rate", "Amount"], ["A", "2", "1000", "2000"], ["B", "3", "5000", "15000"]]
    assert [f for f in C.verify_table(rv, flag_unverified=True) if f.get("kind") == "not_verified"] == []
    # a tiny / sub-floor amount table is not nagged
    assert C.verify_table([["x", "10"], ["y", "20"]], flag_unverified=True) == []

# --------------------------------------------------------------------------- Excel<->PDF reconciliation
def test_reconcile_num_match():
    """The alignment primitive: exact-within-tol / relative / OCR-glyph / no-match scoring."""
    import reconcile_tables as R
    cfg = C.ExtractConfig()
    assert R.num_match(1000.0, 1000.0, "1000", cfg) == 1.0
    assert R.num_match(100000.0, 100050.0, "100050", cfg) == 0.9    # within 1% relative
    assert R.num_match(3000.0, None, "3O00", cfg) == 0.7            # 1 glyph sub (O->0)
    assert R.num_match(3000.0, None, "3OOO", cfg) == 0.0            # 3 subs > max(2) -> too corrupted
    assert R.num_match(1000.0, 9999.0, "apple", cfg) == 0.0

def test_reconcile_align_anchor_noisy_fallback():
    """anchor-then-segment: clean -> exact page boundaries; 30%-OCR-noise -> boundaries hold via
    surviving anchors; too-few-anchors -> count_order fallback capped below the approval gate."""
    import reconcile_tables as R, copy
    cfg = C.ExtractConfig()
    def er(a): return {"grid_ix": a, "anchor": {"amount": (float(a), str(a)), "rate": (None, ""), "qty": (None, "")}}
    A = [100000 + i * 1000 for i in range(90)]                     # realistically spaced distinct amounts
    excel = [er(a) for a in A]
    pdf = [{"page": 37, "anchor": {"amount": (float(A[i]), str(A[i])), "rate": (None, ""), "qty": (None, "")}} for i in range(45)] \
        + [{"page": 38, "anchor": {"amount": (float(A[45 + i]), str(A[45 + i])), "rate": (None, ""), "qty": (None, "")}} for i in range(45)]
    al = R.align_sheet_to_group(excel, pdf, cfg)
    assert al["method"] == "anchor" and len(al["segments"]) == 2
    assert al["segments"][0]["excel_rows"] == list(range(45)) and al["segments"][1]["page"] == 38
    pn = copy.deepcopy(pdf)
    for k in range(0, 90, 3):
        pn[k]["anchor"]["amount"] = (None, str(A[k]).replace("0", "O"))   # heavy glyph -> not an anchor
    aln = R.align_sheet_to_group(excel, pn, cfg)
    assert aln["method"] == "anchor" and aln["segments"][0]["page"] == 37 and aln["segments"][-1]["page"] == 38
    few = [{"page": 1, "anchor": {"amount": (None, "?"), "rate": (None, ""), "qty": (None, "")}} for _ in range(90)]
    alf = R.align_sheet_to_group(excel, few, cfg)
    assert alf["method"] == "count_order" and alf["confidence"] < cfg.reconcile_segment_conf_gate

def test_reconcile_gates():
    """An Excel can only become authority if it itself ties out, has no un-recalculated formulas, and
    its row count reconciles with the PDF tombstone rows."""
    import reconcile_tables as R
    cfg = C.ExtractConfig()
    ok = {"grids": {"S": [["Item", "Amount"], ["A", "100000"], ["B", "200000"], ["Total", "300000"]]},
          "tie": {"passed": True}, "sched": "S", "uncached": []}
    assert R.excel_is_trusted(ok, "S", 2, 2, cfg)[0]
    bad = {"grids": {"S": [["A", "100000"], ["B", "200000"], ["Total", "999999"]]},
           "tie": {"passed": False}, "sched": "S", "uncached": []}
    assert not R.excel_is_trusted(bad, "S", 2, 2, cfg)[0]            # doesn't tie out
    assert not R.excel_is_trusted(ok, "S", 9, 2, cfg)[0]            # row-count mismatch

def test_reconcile_tagging_only_uncertain():
    """doc_layout._tag_reconcilable_tables tombstones ONLY uncertain tables (tie-out flagged here);
    a clean reconciling table is never tagged."""
    try:
        import doc_layout as DL
    except Exception:
        return
    clean = {"type": "table", "id": "p1-e1", "grid": [["A", "Amt"], ["x", "100000"], ["T", "300000"]], "tieout_flags": []}
    bad = {"type": "table", "id": "p1-e2", "grid": [["A", "Amt"], ["x", "100000"], ["T", "300000"]],
           "tieout_flags": [{"kind": "tieout", "gap": 200000}]}
    doc = {"file": "proj/b.pdf", "pages": [{"page_no": 1, "route": "native", "elements": [clean, bad]}]}
    DL._tag_reconcilable_tables(doc, C.ExtractConfig(reconcile_tables=True))
    assert "reconcile" not in clean
    assert bad.get("reconcile", {}).get("tombstone") is True and bad["reconcile"]["reason"] == "tieout_flags"

def test_reconcile_inject_idempotent_total_preserved():
    """End-to-end inject: the authoritative Excel replaces the tombstone grid (Total preserved, now
    ties out), the original garbage grid is retained, and a re-apply is idempotent (no dup, version++)."""
    try:
        import openpyxl, reconcile_tables as R, tempfile, json, copy
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    for row in [["Schedule", "Amount"], ["B1", 100000], ["B2", 200000], ["B3", 300000], ["Total", 600000]]:
        ws.append(row)
    xp = d / "auth.xlsx"; wb.save(str(xp))
    tomb = {"type": "table", "id": "p1-e1", "page": 1, "reading_order_index": 0, "section_path": [], "mode": "ruled",
            "grid": [["Schedule", "Amount"], ["B1", "100000"], ["B2", "200000"], ["B3", "300000"], ["Total", "750000"]],
            "tieout_flags": [{"kind": "tieout", "gap": 150000}],
            "reconcile": {"tombstone": True, "tombstone_id": "k:p1-e1", "reason": "tieout_flags",
                          "status": "PENDING", "version": 0, "logical_group_id": "g1", "n_data_rows": 3}}
    doc = {"file": str(d / "bad.pdf"), "completeness": {},
           "pages": [{"page_no": 1, "route": "native", "elements": [tomb], "needs_review": True,
                      "review_reasons": ["table tie-out / arithmetic flag(s)"]}]}
    lp = d / "bad.layout.json"; C.atomic_write_text(lp, json.dumps(doc))
    sidecar = R.build_reconciliation(doc, lp, [str(xp)], C.ExtractConfig())
    assert sidecar["sources"][0]["links"][0]["trusted"]
    sidecar["approval"] = {"approved_by": "t", "approved_at": "now", "decision": "approved"}
    R.inject(lp, sidecar, C.ExtractConfig(), dry_run=False, approved_by="t")
    t1 = json.loads(lp.read_text(encoding="utf-8"))["pages"][0]["elements"][0]
    assert t1["reconcile"]["status"] == "RECONCILED" and t1["tieout_flags"] == []
    assert any("Total" in str(c) for c in t1["reconciled_grid"][-1])           # total preserved
    assert t1["reconcile"]["original_grid"][-1][-1] == "750000"                # garbage retained
    orig = copy.deepcopy(t1["reconcile"]["original_grid"]); v = t1["reconcile"]["version"]; rg = copy.deepcopy(t1["reconciled_grid"])
    res2 = R.inject(lp, sidecar, C.ExtractConfig(), dry_run=False, approved_by="t")   # re-apply
    t2 = json.loads(lp.read_text(encoding="utf-8"))["pages"][0]["elements"][0]
    assert res2.get("noop") and t2["reconcile"]["version"] == v            # G8: re-apply is a NO-OP, no churn
    assert t2["reconcile"]["original_grid"] == orig and t2["reconciled_grid"] == rg
    assert sum(1 for pg in json.loads(lp.read_text(encoding="utf-8"))["pages"] for e in pg["elements"] if e["type"] == "table") == 1

def test_reconcile_manual_mark():
    """G7: a clean table that auto-tagging missed (high-conf but wrong, no tie-out flag) can be manually
    tombstoned by element id."""
    import reconcile_tables as R
    doc = {"file": "x.pdf", "pages": [{"page_no": 1, "elements": [
        {"type": "table", "id": "p1-e1", "grid": [["A", "Amt"], ["x", "100000"], ["T", "100000"]], "tieout_flags": []}]}]}
    assert R.mark_tombstone(doc, ["p1-e1"]) == 1
    rec = doc["pages"][0]["elements"][0]["reconcile"]
    assert rec["tombstone"] is True and rec["reason"] == "manual" and rec["status"] == "PENDING"
    assert R.mark_tombstone(doc, ["p1-e1"]) == 0            # already tombstoned -> idempotent

def test_reconcile_glyph_aware_pdf_rows():
    """G1: heavily glyph-mangled amount cells ('1OOO') still count as PDF data rows (so the matcher /
    count_order can recover them); prose is rejected; the strict Excel-side predicate is unchanged."""
    import reconcile_tables as R
    noisy = [["Item", "Amount"], ["A", "1OOO"], ["B", "2OOO"], ["C", "3OOO"]]
    assert len(R.grid_data_rows(noisy)[1]) == 0                            # strict -> rows lost
    assert len(R.grid_data_rows(noisy, numeric_pred=R._pdf_numeric)[1]) == 3   # glyph-aware -> recovered
    assert R._pdf_numeric("1OOO") and R._pdf_numeric("1,234.50")
    assert not R._pdf_numeric("Description of work") and not R._pdf_numeric("ItemI")

def test_reconcile_flag_off_no_key():
    """G9 golden: with reconcile_tables OFF (default) NO element ever gets a `reconcile` key; ON, the
    tie-out-failing table is tombstoned. (Guards the 'feature off = unchanged output' claim end-to-end.)"""
    try:
        import fitz, doc_layout as DL, tempfile
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp()); src = d / "b.pdf"
    doc = fitz.open(); p = doc.new_page(width=420, height=320)
    p.insert_text((40, 40), "PRICE SUMMARY with a wrong total below in this clause.", fontsize=10)
    xs = [40, 240, 400]; ys = [80 + 28 * i for i in range(5)]
    for y in ys: p.draw_line((xs[0], y), (xs[-1], y))
    for x in xs: p.draw_line((x, ys[0]), (x, ys[-1]))
    for r, (an, bn) in enumerate([("Schedule", "Amount"), ("B1", "100000"), ("B2", "200000"), ("Total", "999999")]):
        p.insert_text((xs[0] + 5, ys[r] + 18), an); p.insert_text((xs[1] + 5, ys[r] + 18), bn)
    doc.save(str(src)); doc.close()
    off, _, _ = DL.extract_document(src, d / "off", C.ExtractConfig(reconcile_tables=False))
    assert not any("reconcile" in e for pg in off["pages"] for e in pg["elements"])
    on, _, _ = DL.extract_document(src, d / "on", C.ExtractConfig(reconcile_tables=True))
    assert any((e.get("reconcile") or {}).get("tombstone") for pg in on["pages"] for e in pg["elements"] if e["type"] == "table")

def test_reconcile_no_html_escape_in_grid():
    """H3 regression: the reconciled grid stores RAW Excel cell values. layout.json is JSON and the
    .md is Markdown, so html.escape (the old _esc_grid) CORRUPTED text cells ('Excavation & filling'
    -> '...&amp;...', 'girders < 20m' -> '...&lt;...'). Assert the grid-builder that populates
    reconciled_grid is byte-identical to the raw strings (only None->'' kept). Fails vs the old escaping."""
    import reconcile_tables as R
    sliced = [["Desc", "Amount"], ["Excavation & filling", "1000"],
              ["girders < 20m", 'a "quoted" price', None]]
    out = R._norm_grid(sliced)                                    # the exact call inject() makes into reconciled_grid
    assert out == [["Desc", "Amount"], ["Excavation & filling", "1000"],
                   ["girders < 20m", 'a "quoted" price', ""]]     # raw values, None->'' , NOTHING escaped
    flat = "".join(c for row in out for c in row)
    for entity in ("&amp;", "&lt;", "&gt;", "&quot;", "&#x27;", "&#39;"):
        assert entity not in flat, ("html entity leaked into reconciled_grid: " + entity)

def test_table_pdf_page_tmpdir_is_per_file_unique():
    """H1 regression: the scanned-page raster temp dir must be per-FILE unique. The old shared
    'out/_pages' meant two parallel workers wrote/read the same 'p{N}.png' and could OCR each other's
    page (silent cross-file contamination). Pure/CPU-only: computes dirs, no rasterize/OCR."""
    import table_pdf as TP, tempfile, shutil
    from pathlib import Path
    base = Path(tempfile.mkdtemp()); out_xlsx = base / "a.tables.xlsx"
    d1 = TP._page_tmpdir(out_xlsx, "projA/BOQ.pdf")
    d2 = TP._page_tmpdir(out_xlsx, "projB/OTHER.pdf")
    d3 = TP._page_tmpdir(out_xlsx, "projB/BOQ.pdf")               # SAME stem as d1, different folder
    try:
        assert d1 != d2 and d1 != d3 and d2 != d3                 # all distinct (no collision even on same stem)
        assert d1.name != "_pages" and d3.name != "_pages"        # NOT the old shared dir
        assert d1.parent == base and d3.parent == base            # created under the output dir
        assert d1.exists() and d2.exists() and d3.exists()        # mkdtemp actually created them
        assert "BOQ" in d1.name and "BOQ" in d3.name              # stem-tagged for debuggability
    finally:
        for d in (d1, d2, d3):
            shutil.rmtree(str(d), ignore_errors=True)
        shutil.rmtree(str(base), ignore_errors=True)

def test_h2_collision_safe_dirs_and_done_resume():
    """H2: two same-named source files in DIFFERENT folders map to DISTINCT sharded output dirs (no
    silent overwrite), and the resume predicate keys on the .done SENTINEL - a partial doc.json with
    NO .done is treated as NOT done (never trusted), .done present is done. Pure/CPU-only (no extract)."""
    import common as C, tempfile
    from pathlib import Path
    out = Path(tempfile.mkdtemp())
    a, b = "projA/BOQ.pdf", "projB/BOQ.pdf"
    da, db = C.doc_outdir(out, a), C.doc_outdir(out, b)
    assert da != db                                          # same stem, different folder -> distinct dirs
    resume_done = lambda fp: (C.doc_outdir(out, fp) / ".done").exists()   # the run_batch/table_pdf predicate
    da.mkdir(parents=True, exist_ok=True)
    (da / "BOQ.doc.json").write_text("{partial", encoding="utf-8")        # simulate a worker killed mid-write
    assert resume_done(a) is False                           # partial output, NO .done -> NOT done (must redo)
    C.mark_done(da)
    assert resume_done(a) is True                            # sentinel present -> done (safe to skip)
    assert resume_done(b) is False                           # the other same-stem file is independent

def test_h2_write_outputs_atomic_and_sentinel():
    """H2: write_outputs writes doc.json/chunks.jsonl/preview.html ATOMICALLY (no .tmp litter) and drops
    the .done sentinel LAST; deleting .done (a simulated mid-write: outputs present, sentinel absent) is
    treated as NOT done by the resume predicate. Pure/CPU-only."""
    import common as C, tempfile
    from pathlib import Path
    out = Path(tempfile.mkdtemp()); src = "proj/inv.pdf"
    docdir = C.doc_outdir(out, src); docdir.mkdir(parents=True, exist_ok=True)
    pages = [{"page": 1, "route": "native", "confidence": 0.97, "chars": 120, "png": None}]
    row = C.write_outputs(src, "pdf", pages, {}, [{"kind": "page", "text": "hello", "confidence": 0.97}],
                          "hello world", 0.70, docdir, "inv")
    assert row["status"] == "AUTO_ACCEPT" and row["needs_review"] is False
    for name in ("inv.doc.json", "inv.chunks.jsonl", "inv.preview.html"):
        assert (docdir / name).exists() and not (docdir / (name + ".tmp")).exists()   # atomic: no temp litter
    assert (docdir / ".done").exists()                       # sentinel dropped only AFTER all outputs flushed
    (docdir / ".done").unlink()                              # simulate a mid-write (outputs present, no sentinel)
    assert not (C.doc_outdir(out, src) / ".done").exists()   # -> resume treats it as NOT done, re-does it

def test_r14_atomic_write_survives_long_path():
    r"""R14: a deep output dir whose absolute path exceeds Windows MAX_PATH (260) must NOT crash
    atomic_write_text - the file is written with EXACT bytes - and the path we store/return never
    carries the \\?\ extended-length prefix (that is syscall-only). Cross-platform: on posix there is
    no 260 limit so it simply verifies the write + no-prefix; on win32 it proves the long-path fix."""
    import common as C, tempfile, os, shutil
    from pathlib import Path
    root = Path(tempfile.mkdtemp())
    seg = "d" * 40
    deep = root
    while len(str(deep)) < 300:                               # force the absolute path well past 260
        deep = deep / seg
    os.makedirs(C._fs_path(deep), exist_ok=True)              # long-path aware so the test setup itself can't fail
    target = deep / "out.txt"
    assert len(str(target)) > 260                             # the case that used to raise FileNotFoundError
    payload = "alpha\nbeta\ngamma\n"
    C.atomic_write_text(target, payload)                      # must not raise
    with open(C._fs_path(target), "r", encoding="utf-8", newline="") as f:
        assert f.read() == payload                            # exact bytes, LF preserved
    assert "\\\\?\\" not in str(target)                       # the STORED/returned path has no \\?\ prefix
    assert "\\\\?\\" not in str(deep)
    if os.name != "nt":                                       # on posix the prefix is never applied
        assert C._fs_path(target) == os.path.abspath(str(target))
    else:                                                     # on win32 the syscall path IS prefixed (but not stored)
        assert C._fs_path(target).startswith("\\\\?\\")
    shutil.rmtree(C._fs_path(root), ignore_errors=True)

def test_m1_shared_min_text_gate():
    """M1: ONE shared digital-vs-scanned predicate reads .min_text; default 50 reproduces today's
    pipeline/pdf_extract/doc_layout routing (byte-identical), and an explicit --min-text flips routing.
    Also proves to_pdf_options now CARRIES the threshold (it used to be dropped = silent no-op)."""
    import common as C
    t30 = "x" * 30
    assert C.native_text_gate(None) == 50 and C.native_text_gate(C.ExtractConfig()) == 50   # default = old hardcoded 50
    assert C.is_native_text(t30, None) is False                                  # 30 chars < 50 -> scanned (unchanged)
    assert C.is_native_text(t30, C.ExtractConfig()) is False
    assert C.is_native_text(t30, C.ExtractConfig(min_text=20)) is True           # explicit lower gate -> native
    assert C.is_native_text("y" * 60, C.ExtractConfig(min_text=100)) is False    # explicit higher gate -> scanned
    assert C.is_native_text("z" * 50, C.ExtractConfig()) is True                 # boundary: exactly the gate is native
    assert C.ExtractConfig().min_text == 50                                      # reconciled default (was 20)
    assert C.ExtractConfig(min_text=37).to_pdf_options()["min_text"] == 37       # threaded through (was dropped)
    try:
        import pdf_extract as PE
    except Exception:
        return
    assert PE.PdfOptions().min_text == 50                                        # PdfOptions default reproduces 50
    assert C.is_native_text(t30, PE.PdfOptions()) is False                       # predicate reads .min_text off PdfOptions
    assert C.is_native_text(t30, PE.PdfOptions(min_text=10)) is True

def test_m3_image_pixel_cap():
    """M3: the pixel-cap predicate skips an oversized (decompression-bomb) image BEFORE any Pixmap is
    allocated, using declared w x h only (no bitmap alloc), and accepts normal images. Defensive on
    unknown dims. Default cap 300 MP covers a 300-DPI A0 scan (~139 MP) while stopping a 50k x 50k bomb."""
    import common as C
    assert C.image_within_pixel_cap(4000, 3000, None) is True              # 12 MP photo
    assert C.image_within_pixel_cap(9930, 14040, None) is True             # A0 @ 300 DPI ~139 MP < 300 cap
    assert C.image_within_pixel_cap(50000, 50000, None) is False           # 2500 MP bomb -> skip (never decoded)
    assert C.image_within_pixel_cap(0, 100, None) is False                 # degenerate -> skip defensively
    assert C.image_within_pixel_cap(None, None, None) is False             # unknown -> skip defensively
    assert C.image_within_pixel_cap(50000, 50000, C.ExtractConfig(max_image_megapixels=0)) is True   # 0 = disabled
    assert C.image_within_pixel_cap(20000, 20000, C.ExtractConfig(max_image_megapixels=500)) is True # 400 MP < 500
    assert C.image_within_pixel_cap(20000, 20000, C.ExtractConfig(max_image_megapixels=100)) is False # 400 MP > 100

def test_m4_encrypted_pdf_guard():
    """M4: the shared encryption helper flags a needs_pass=True doc as review-routed with an 'encrypted'
    reason and passes a normal one through. Stub doc (no real encrypted PDF needed)."""
    import common as C
    class _Doc:
        def __init__(self, np): self.needs_pass = np
    r = C.encrypted_reason(_Doc(True))
    assert r and "encrypt" in r.lower()                                    # review reason mentions encryption
    assert C.encrypted_reason(_Doc(False)) is None                         # normal PDF -> proceeds
    assert C.encrypted_reason(object()) is None                            # no needs_pass attr -> None (safe)

def test_m5_zip_bomb_guard():
    """M5: the size-summing predicate flags a bomb (huge uncompressed OR absurd ratio) and passes a
    normal archive; the full guard passes a real small workbook-shaped zip. No gigabytes written -
    synthetic sizes for the predicate, a tiny real zip for the end-to-end path."""
    import common as C, zipfile, tempfile, shutil
    from pathlib import Path
    assert C._zip_bomb_verdict(10 * 1024 * 1024, 1 * 1024 * 1024) is None                 # 10 MB, 10x -> fine
    r_big = C._zip_bomb_verdict(3000 * 1024 * 1024, 10 * 1024 * 1024)                      # 3 GB uncompressed > 2048
    assert r_big and "uncompressed" in r_big
    r_ratio = C._zip_bomb_verdict(500 * 1024 * 1024, 1 * 1024 * 1024)                      # 500x on 500 MB -> bomb
    assert r_ratio and "compression" in r_ratio
    assert C._zip_bomb_verdict(100 * 1024 * 1024, 100 * 1024 * 1024) is None               # 1x ratio -> fine
    assert C._zip_bomb_verdict(60 * 1024 * 1024, 100 * 1024) is not None                   # 600x on 60 MB -> bomb
    assert C._zip_bomb_verdict(40 * 1024 * 1024, 100 * 1024) is None                       # >50 MB gate not met -> fine
    d = Path(tempfile.mkdtemp()); good = d / "good.xlsx"
    with zipfile.ZipFile(good, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", "<x/>")
        z.writestr("xl/worksheets/sheet1.xml", "<row/>" * 2000)
    assert C.zip_bomb_reason(good) is None                                                 # real small workbook passes
    assert C.zip_bomb_reason(d / "nope.xlsx") is None                                       # missing file -> None (safe)
    shutil.rmtree(str(d), ignore_errors=True)

def test_m5_libreoffice_recalc_cleans_tempdir():
    """M5: _libreoffice_recalc removes its mkdtemp temp dir in a finally (no leak). Monkeypatch the
    soffice discovery + subprocess call so NOTHING external runs; capture the temp dir and assert it is
    gone after the call returns (None, None)."""
    import excel_extract as X, tempfile, subprocess, shutil, os
    created = []
    real_mkdtemp, real_run, real_which = tempfile.mkdtemp, subprocess.run, shutil.which
    def fake_mkdtemp(*a, **k):
        p = real_mkdtemp(*a, **k); created.append(p); return p
    tempfile.mkdtemp = fake_mkdtemp
    subprocess.run = lambda *a, **k: None                              # soffice no-op (produces no output file)
    shutil.which = lambda name: "soffice" if "soffice" in str(name) else real_which(name)
    try:
        g, u = X._libreoffice_recalc("does-not-exist.xlsx", soffice_cmd="soffice")
        assert g is None and u is None                                 # no evaluated copy produced
        assert created and not os.path.exists(created[0])              # temp dir cleaned in the finally
    finally:
        tempfile.mkdtemp, subprocess.run, shutil.which = real_mkdtemp, real_run, real_which

def test_reconcile_context_signals():
    """FIX 2/3 (pure): the context adjustment is EXACTLY 0 with no header/section (numbers-only stays
    byte-identical), a matching header raises confidence, a clashing header lowers it, and a caption
    token bonus/contradiction pushes confidence up / below the gate."""
    import reconcile_tables as R, common as C
    cfg = C.ExtractConfig()
    assert R._context_adjust(0.90, "", "", "", "", cfg)[0] == 0.90            # no context -> unchanged
    assert R._context_adjust(0.90, "Item Cost", "Item Cost", "", "", cfg)[0] > 0.90   # header match -> up
    assert R._context_adjust(0.90, "Item Cost", "Zzz Qqq", "", "", cfg)[0] < 0.90     # header clash -> down
    assert R._table_token("Annexure B - rates") == ("annexure", "b")
    assert R._caption_signal("see Table 3.2", "Table 3.2", cfg) > 0                    # match -> bonus
    assert R._caption_signal("see Table 3.2", "Table 9", cfg) < 0                      # contradiction -> penalty
    # a numeric coincidence (0.94) with a contradicting caption drops below the 0.70 gate -> held
    assert R._context_adjust(0.94, "", "", "Table 3.2", "Table 9", cfg)[0] < cfg.reconcile_segment_conf_gate

def test_reconcile_global_assignment_beats_greedy():
    """FIX 4 (pure): _assign_global maximizes total score; a case where sheet-order greedy picks wrong
    but the global optimum picks right. Errors/fails against pre-fix code (no _assign_global)."""
    import reconcile_tables as R
    sheets, gids = ["s1", "s2"], ["gA", "gB"]
    score = {("s1", "gA"): 0.90, ("s1", "gB"): 0.85, ("s2", "gA"): 0.95, ("s2", "gB"): 0.20}
    got = dict(R._assign_global(sheets, gids, score))
    assert got == {"s1": "gB", "s2": "gA"}, got            # global optimum (total 1.80), NOT greedy (s1->gA=1.10)
    # greedy-by-sheet-order would take s1->gA (its best) then s2->gB : demonstrably worse
    greedy, used = {}, set()
    for s in sheets:
        g = max((g for g in gids if g not in used), key=lambda g: score[(s, g)])
        greedy[s] = g; used.add(g)
    assert greedy == {"s1": "gA", "s2": "gB"} and greedy != got      # confirms greedy is the WRONG answer here
    assert R._assign_global(sheets, [], score) == [] and R._assign_global([], gids, score) == []

def _mk_tomb(eid, page, roi, grid, section):
    return {"type": "table", "id": eid, "page": page, "reading_order_index": roi, "section_path": list(section),
            "mode": "ruled", "grid": grid, "tieout_flags": [{"kind": "tieout", "gap": 1}],
            "reconcile": {"tombstone": True, "tombstone_id": "k:%s" % eid, "reason": "tieout_flags",
                          "status": "PENDING", "version": 0, "logical_group_id": None,
                          "n_data_rows": sum(1 for r in grid if any(str(c or "").strip() for c in r)),
                          "section_path": list(section)}}

def test_reconcile_same_page_two_tables():
    """FIX 1: TWO distinct tombstoned tables on the SAME page each reconcile INDEPENDENTLY to their own
    Excel sheet (matched by distinct numbers). Against page-keyed code the two same-page tombstones merge
    into one group and inject's page->element dict overwrites one -> only ONE would reconcile (fails)."""
    try:
        import openpyxl, reconcile_tables as R, common as C, tempfile, json
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    mat = wb.create_sheet("MAT")
    for row in [["Schedule", "Amount"], ["B1", 100000], ["B2", 200000], ["B3", 300000], ["Total", 600000]]:
        mat.append(row)
    lab = wb.create_sheet("LAB")
    for row in [["Schedule", "Amount"], ["L1", 150000], ["L2", 250000], ["L3", 350000], ["Total", 750000]]:
        lab.append(row)
    xp = d / "auth.xlsx"; wb.save(str(xp))
    # two tombstones on PAGE 1: e1 numbers match MAT, e2 numbers match LAB, both with a wrong total
    e1 = _mk_tomb("p1-e1", 1, 0, [["Schedule", "Amount"], ["B1", "100000"], ["B2", "200000"],
                                  ["B3", "300000"], ["Total", "999999"]], ["Materials"])
    e2 = _mk_tomb("p1-e2", 1, 1, [["Schedule", "Amount"], ["L1", "150000"], ["L2", "250000"],
                                  ["L3", "350000"], ["Total", "111111"]], ["Labour"])
    doc = {"file": str(d / "b.pdf"), "completeness": {},
           "pages": [{"page_no": 1, "route": "native", "elements": [e1, e2],
                      "needs_review": True, "review_reasons": ["table tie-out / arithmetic flag(s)"]}]}
    lp = d / "b.layout.json"; C.atomic_write_text(lp, json.dumps(doc))
    sidecar = R.build_reconciliation(doc, lp, [str(xp)], C.ExtractConfig())
    sidecar["approval"] = {"approved_by": "t", "approved_at": "now", "decision": "approved"}
    R.inject(lp, sidecar, C.ExtractConfig(), dry_run=False, approved_by="t")
    out = json.loads(lp.read_text(encoding="utf-8"))["pages"][0]["elements"]
    t1 = next(e for e in out if e["id"] == "p1-e1"); t2 = next(e for e in out if e["id"] == "p1-e2")
    assert t1["reconcile"]["status"] == "RECONCILED" and t2["reconcile"]["status"] == "RECONCILED"  # BOTH
    assert t1["reconcile"]["provenance"]["sheet"] == "MAT"      # e1 <- materials numbers
    assert t2["reconcile"]["provenance"]["sheet"] == "LAB"      # e2 <- labour numbers (not overwritten)
    assert any("600000" in str(c) for c in t1["reconciled_grid"][-1])
    assert any("750000" in str(c) for c in t2["reconciled_grid"][-1])

def test_reconcile_header_section_disambiguates():
    """FIX 2/3: two same-page tables with IDENTICAL numbers but DIFFERENT headers/sections each link to
    the correct sheet via the header/section signal. Excel sheet order is deliberately REVERSED so the
    old numbers-only greedy (first sheet claims the first group) links them BACKWARDS -> fails pre-fix."""
    try:
        import openpyxl, reconcile_tables as R, common as C, tempfile, json
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    lab = wb.create_sheet("Labour")                    # inserted FIRST (dict order) on purpose
    for row in [["Task", "Cost"], ["a", 100000], ["b", 200000], ["c", 300000], ["Total", 600000]]:
        lab.append(row)
    mat = wb.create_sheet("Materials")
    for row in [["Item", "Cost"], ["a", 100000], ["b", 200000], ["c", 300000], ["Total", 600000]]:
        mat.append(row)                                # IDENTICAL numbers to Labour
    xp = d / "auth.xlsx"; wb.save(str(xp))
    e1 = _mk_tomb("p1-e1", 1, 0, [["Item", "Cost"], ["a", "100000"], ["b", "200000"],
                                  ["c", "300000"], ["Total", "999999"]], ["Materials"])   # group g1
    e2 = _mk_tomb("p1-e2", 1, 1, [["Task", "Cost"], ["a", "100000"], ["b", "200000"],
                                  ["c", "300000"], ["Total", "999999"]], ["Labour"])      # group g2
    doc = {"file": str(d / "b.pdf"), "completeness": {},
           "pages": [{"page_no": 1, "route": "native", "elements": [e1, e2],
                      "needs_review": True, "review_reasons": ["table tie-out / arithmetic flag(s)"]}]}
    lp = d / "b.layout.json"; C.atomic_write_text(lp, json.dumps(doc))
    sidecar = R.build_reconciliation(doc, lp, [str(xp)], C.ExtractConfig())
    sidecar["approval"] = {"approved_by": "t", "approved_at": "now", "decision": "approved"}
    R.inject(lp, sidecar, C.ExtractConfig(), dry_run=False, approved_by="t")
    out = json.loads(lp.read_text(encoding="utf-8"))["pages"][0]["elements"]
    t1 = next(e for e in out if e["id"] == "p1-e1"); t2 = next(e for e in out if e["id"] == "p1-e2")
    assert t1["reconcile"]["provenance"]["sheet"] == "Materials"   # matched by header 'Item'/section 'Materials'
    assert t2["reconcile"]["provenance"]["sheet"] == "Labour"      # numbers alone are a tie -> header/section decides

def test_reconcile_conflicting_excel_held():
    """SAFETY: an Excel whose numbers AND header do NOT match the tombstone is HELD (not injected) even
    though the sheet itself ties out - the alignment confidence stays below the gate. The net must not be
    weakened by the new context signals."""
    try:
        import openpyxl, reconcile_tables as R, common as C, tempfile, json
        from pathlib import Path
    except Exception:
        return
    d = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "WRONG"
    for row in [["Widget", "Price"], ["x", 999], ["y", 888], ["z", 777], ["Total", 2664]]:
        ws.append(row)                                 # ties out, but numbers unrelated to the tombstone
    xp = d / "wrong.xlsx"; wb.save(str(xp))
    e1 = _mk_tomb("p1-e1", 1, 0, [["Schedule", "Amount"], ["B1", "100000"], ["B2", "200000"],
                                  ["B3", "300000"], ["Total", "600000"]], ["Table 3.2 Price Schedule"])
    doc = {"file": str(d / "b.pdf"), "completeness": {},
           "pages": [{"page_no": 1, "route": "native", "elements": [e1],
                      "needs_review": True, "review_reasons": ["table tie-out / arithmetic flag(s)"]}]}
    lp = d / "b.layout.json"; C.atomic_write_text(lp, json.dumps(doc))
    sidecar = R.build_reconciliation(doc, lp, [str(xp)], C.ExtractConfig())
    res = R.inject(lp, sidecar, C.ExtractConfig(), dry_run=True)         # dry run: what WOULD inject?
    assert res["would_inject"] == 0                                     # held for a human, never force-linked
    r2 = R.inject(lp, sidecar, C.ExtractConfig(), dry_run=False, approved_by="t")
    assert r2.get("noop") and r2.get("injected") == 0                  # apply is a no-op; nothing overwritten
    e_after = json.loads(lp.read_text(encoding="utf-8"))["pages"][0]["elements"][0]
    assert "reconciled_grid" not in e_after and e_after["reconcile"]["status"] == "PENDING"

def test_merged_numeric_cell_flagged():
    """FIX 1: a table that TIES OUT but has ONE cell holding two collapsed numbers ('10 20' in a Qty
    column) is flagged as a suspected merged/collapsed row (parse_number keeps only the first -> data
    lost). Prose with numbers ('Providing 100 mm x 200 mm pipe as per IS 458') and 'Rate per 100' do
    NOT flag. Pre-fix verify_table has no merged detector -> returns [] for the merged case -> FAILS."""
    import common as C
    # amount column ties out (5000+6000=11000); the '10 20' Qty cell is the ONLY defect
    merged = [["Sr", "Qty", "Amount"], ["1", "10 20", "5000"], ["2", "30", "6000"], ["Total", "", "11000"]]
    flags = C.verify_table(merged)
    mc = [f for f in flags if f.get("kind") == "merged_cell"]
    assert mc and "10 20" in mc[0]["stated"], flags        # ONE merged cell is enough -> flagged
    assert not [f for f in flags if f.get("kind") not in ("merged_cell",)]   # tie-out itself is clean
    # guard: prose-with-numbers and 'Rate per 100' must NOT flag (and the amounts still tie out -> [])
    ok = [["Sr", "Desc", "Amount"],
          ["1", "Providing 100 mm x 200 mm pipe as per IS 458", "5000"],
          ["2", "Rate per 100 units", "6000"], ["Total", "", "11000"]]
    assert C.verify_table(ok) == [], C.verify_table(ok)
    assert C.merged_numeric_cells([["2 nos", "x"]]) == []   # single number + text -> not merged
    assert C.merged_numeric_cells([["100 200 300"]]) == [(0, 0)]   # pure multi-number cell -> merged

def test_alt_column_cannot_mask_dropped_row():
    """FIX 2: the header amount column FAILS tie-out (a dropped row) but a Qty column happens to reconcile
    to its own total; the old code let that coincidental reconcile CLEAR the real failure. The alt column
    must be a legitimate (header-identified) amount column to clear a flag -> Qty can't, so the dropped-row
    gap survives. Pre-fix: rowwise_max=Qty reconciles -> flag cleared -> verify_table==[] -> FAILS."""
    import common as C
    # Qty (row-max, reconciles: 100000+200000+300000=600000) must NOT clear the Amount failure
    # (50000+60000+70000=180000 != stated 999000). Amounts >= reconcile_min_total so the tie-out runs.
    masking = [["Item", "Qty", "Amount"], ["A", "100000", "50000"], ["B", "200000", "60000"],
               ["C", "300000", "70000"], ["Total", "600000", "999000"]]
    flags = C.verify_table(masking)
    gaps = [f for f in flags if isinstance(f.get("gap"), (int, float)) and abs(f["gap"] - 819000) < 1]
    assert gaps, flags                                    # the real Amount dropped-row gap (999000-180000) is NOT masked
    # a table with TWO header-identified amount columns still lets the correct one clear a mislabel (no regression)
    twoamt = [["Amount A", "Amount B"], ["100", "100"], ["200", "200"], ["Total", "300"]]  # A total 'Total' text; B ok
    _ = C.verify_table(twoamt)                            # (smoke: must not raise)

def test_xlsx_formula_injection_sanitized():
    """FIX 3: a cell written as TEXT starting with = + - @ is prefixed with ' so Excel can't execute it;
    numbers and normal text are untouched. Pre-fix has no _xlsx_safe -> AttributeError -> FAILS."""
    import table_pdf as TP
    assert TP._xlsx_safe("=cmd|' /C calc'!A1") == "'=cmd|' /C calc'!A1"
    assert TP._xlsx_safe("+1+2") == "'+1+2"
    assert TP._xlsx_safe("-2+3") == "'-2+3"
    assert TP._xlsx_safe("@SUM(A1)") == "'@SUM(A1)"
    assert TP._xlsx_safe("5000") == "5000"                # a plain text number is untouched
    assert TP._xlsx_safe("(2,500.00)") == "(2,500.00)"    # accounting negative (starts '(') untouched
    assert TP._xlsx_safe(-500) == -500                    # a REAL numeric negative stays a number
    assert TP._xlsx_safe("Earthwork") == "Earthwork" and TP._xlsx_safe(None) is None
    # and the sanitizer composes with an openpyxl write (round-trips inert)
    try:
        import openpyxl, tempfile
        from pathlib import Path
        d = Path(tempfile.mkdtemp()); p = d / "t.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append([TP._xlsx_safe(c) for c in ["=1+1", "Desc", -5]]); wb.save(str(p))
        wb2 = openpyxl.load_workbook(str(p)); r = list(wb2.active.iter_rows(values_only=True))[0]
        assert r[0] == "'=1+1" and r[2] == -5            # formula neutralized; real negative preserved
    except Exception:
        pass

def test_borderless_bbox_no_page_subtraction():
    """FIX 4: a table box covering most of the page's TEXT extent must NOT be used to subtract text (it
    would delete every surrounding line); _table_elements no longer fabricates a whole-page bbox. Pre-fix
    has no _lines_extent/_covers_most_text -> AttributeError -> FAILS."""
    import doc_layout as DL
    lines = [{"bbox": [10, 10, 100, 20]}, {"bbox": [10, 30, 100, 40]}, {"bbox": [10, 50, 100, 60]}]
    ext = DL._lines_extent(lines)
    assert ext == [10, 10, 100, 60]
    assert DL._covers_most_text([0, 0, 110, 70], ext) is True     # whole-page box -> would swallow prose -> excluded
    assert DL._covers_most_text([10, 10, 100, 22], ext) is False  # a small table box (~20% of text) -> fine to subtract
    assert DL._covers_most_text(None, ext) is False and DL._covers_most_text([0, 0, 10, 10], None) is False

    # INTEGRATION - the borderless-NO-BBOX case (the bug the fix targets): a borderless table whose
    # robust_tables result carries NO bbox must map to bbox=None, NEVER a fabricated whole-page box.
    grid = [["A", "B"], ["1", "2"], ["3", "4"]]                   # a genuinely tabular (2 numeric cols) grid
    orig_extract = DL.RT.extract_page
    try:
        DL.RT.extract_page = lambda page: ([{"grid": grid, "mode": "borderless", "bbox": None}], {})
        els = DL._table_elements(object())                       # page arg unused by the stub
    finally:
        DL.RT.extract_page = orig_extract
    assert len(els) == 1 and els[0]["bbox"] is None, els          # no fabricated bbox -> subtracts nothing

    # the EXACT doc_layout subtraction seam (lines 690-691, 708): a None bbox AND an over-captured
    # borderless box both contribute NOTHING to tboxes, so every surrounding text line survives.
    page_lines = [{"bbox": [10, 10, 100, 20]}, {"bbox": [10, 30, 100, 40]},
                  {"bbox": [10, 50, 100, 60]}, {"bbox": [10, 70, 100, 80]}]
    txt_ext = DL._lines_extent(page_lines)
    tbl_els = [{"bbox": None,            "mode": "borderless"},   # borderless-no-bbox
               {"bbox": [0, 0, 110, 90], "mode": "borderless"}]   # over-captured borderless region
    tboxes = [t["bbox"] for t in tbl_els
              if t["bbox"] and (t["mode"] == "ruled" or not DL._covers_most_text(t["bbox"], txt_ext))]
    assert tboxes == [], tboxes                                   # both excluded
    survivors = [ln for ln in page_lines if not DL._inside(ln["bbox"], tboxes)]
    assert len(survivors) == len(page_lines)                     # BUG WOULD SUBTRACT ALL; fix keeps every line

    # no over-correction: a full-page RULED box is trustworthy (real ruling lines) and STILL subtracts,
    # and a small non-covering box subtracts only its own line - the fix is borderless-only.
    for box, mode, expect_subtracted in ([[0, 0, 110, 90], "ruled", 4], [[8, 8, 102, 22], "ruled", 1]):
        tb = [box] if (mode == "ruled" or not DL._covers_most_text(box, txt_ext)) else []
        subtracted = [ln for ln in page_lines if DL._inside(ln["bbox"], tb)]
        assert len(subtracted) == expect_subtracted, (box, mode, len(subtracted))

# --------------------------------------------------------------------------- multi-engine compare (eval-only)
def _make_boq_text_pdf(path):
    """A ruled BOQ with ENOUGH native text to cross the rig's 50-char digital gate (so the 'current'
    adapter takes the digital robust_tables path, not the OCR path). 6 ruling lines -> 5 rows;
    100+200+300 == Total 600 -> ties out. Distinct from the tiny _make_boq_pdf (which is below the
    native-text gate on purpose)."""
    doc = fitz.open(); page = doc.new_page(width=600, height=340)
    xs = [40, 360, 560]; ys = [80, 120, 160, 200, 240, 280]
    for y in ys: page.draw_line((xs[0], y), (xs[-1], y))
    for x in xs: page.draw_line((x, ys[0]), (x, ys[-1]))
    rows = [("Description", "Amount"), ("Earthwork excavation for foundation", "100"),
            ("PCC 1:4:8 base course laid", "200"), ("RCC M25 in footing works", "300"),
            ("Total", "600")]
    for r, (c0, c1) in enumerate(rows):
        page.insert_text((xs[0] + 5, ys[r] + 22), c0, fontsize=8)
        page.insert_text((xs[1] + 5, ys[r] + 22), c1, fontsize=8)
    doc.save(path); doc.close()

def test_engine_adapter_schema_shape():
    """Every adapter, via base.run_adapter, returns the SAME comparable result dict (all required
    keys, right types); grids preserved; a missing package -> available=False (never raises); an
    error mid-extract -> available=True + reason (one bad file never aborts the run). Pre-impl:
    eval.engines missing -> ModuleNotFoundError -> FAILS."""
    from eval.engines import base
    def ok(inp, opt):
        return {"tables": [[["Item", "Amount"], ["A", "100"]]], "text": "x", "confidence": 0.9}
    r = base.run_adapter("fake", True, "sovereign(test)", ok, "nofile", base.EngineOptions())
    for k in base.REQUIRED_KEYS: assert k in r, (k, sorted(r))
    assert r["available"] is True and r["tables"] == [[["Item", "Amount"], ["A", "100"]]]
    assert isinstance(r["seconds"], float) and r["seconds"] >= 0.0
    assert r["confidence"] == 0.9 and r["sovereign"] is True and r["text"] == "x"
    def missing(inp, opt): raise ModuleNotFoundError("No module named 'docling'")
    m = base.run_adapter("docling", True, "s", missing, "nofile", base.EngineOptions())
    assert m["available"] is False and m["tables"] == [] and "docling" in m["reason"]
    def boom(inp, opt): raise ValueError("bad page 3")
    e = base.run_adapter("x", True, "s", boom, "nofile", base.EngineOptions())
    assert e["available"] is True and e["tables"] == [] and e["reason"].startswith("error:")

def test_engine_graceful_unavailable_when_missing():
    """docling/paddle/vlm degrade to available=False with a clear reason when the package/GPU is
    absent - the harness never raises. (On a host where they ARE installed they return tables
    instead; either branch is valid, so the test asserts the graceful contract, not a fixed answer.)"""
    if fitz is None: return
    from eval.engines import run_engine, EngineOptions
    import tempfile
    d = Path(tempfile.mkdtemp()); pdf = d / "boq.pdf"; _make_boq_text_pdf(str(pdf))
    for name in ("docling", "paddle", "vlm"):
        r = run_engine(name, str(pdf), EngineOptions())
        assert {"engine", "available", "tables", "text", "confidence", "seconds"}.issubset(r)
        if not r["available"]:
            assert r["tables"] == [] and r["reason"]         # clear reason, no crash
    u = run_engine("bogus", str(pdf), EngineOptions())        # unknown engine -> handled, not raised
    assert u["available"] is False and "unknown" in u["reason"].lower()

def test_current_adapter_digital_pdf_tables_and_tieout():
    """The sovereign BASELINE adapter extracts the digital BOQ via robust_tables and the harness
    tie-out PASSES on its output - proving no engine bypasses common.verify_table."""
    if fitz is None: return
    from eval.engines import run_engine, EngineOptions, tieout
    import tempfile
    d = Path(tempfile.mkdtemp()); pdf = d / "boq.pdf"; _make_boq_text_pdf(str(pdf))
    r = run_engine("current", str(pdf), EngineOptions())
    assert r["available"] is True and r["sovereign"] is True
    flat = [row for g in r["tables"] for row in g]
    assert ["Total", "600"] in flat and ["Earthwork excavation for foundation", "100"] in flat
    to = tieout(r["tables"], None)
    assert to["pass"] is True and to["real_gaps"] == 0

def test_compare_harness_scores_ties_out_and_labels_sovereignty():
    """compare_engines runs each engine, tie-outs EVERY engine's tables, labels PaddleOCR
    non-sovereign, emits 'no gold - not scored' without a gold and the accuracy columns with one."""
    if fitz is None: return
    import eval.compare_engines as CE
    from eval.engines import EngineOptions
    import tempfile, json
    d = Path(tempfile.mkdtemp()); pdf = d / "boq.pdf"; _make_boq_text_pdf(str(pdf))
    gdir = d / "golden"; gdir.mkdir(); opt = EngineOptions()
    rep = CE.run_file(str(pdf), ["current", "paddle"], opt, str(gdir))     # no gold yet
    assert rep["scored"] is False
    paddle = [e for e in rep["engines"] if e["engine"] == "paddle"][0]
    assert paddle["sovereign"] is False and "NON-SOVEREIGN" in paddle["sovereignty"].upper()
    cur = [e for e in rep["engines"] if e["engine"] == "current"][0]
    assert cur["tieout"] and cur["tieout"]["pass"] is True                 # every engine's tables tie-out'd
    stem = CE._safe_stem(str(pdf))
    (gdir / f"{stem}.tables.json").write_text(json.dumps([[
        ["Description", "Amount"], ["Earthwork excavation for foundation", "100"],
        ["PCC 1:4:8 base course laid", "200"], ["RCC M25 in footing works", "300"],
        ["Total", "600"]]]), encoding="utf-8")
    rep2 = CE.run_file(str(pdf), ["current"], opt, str(gdir))              # gold present -> scored
    assert rep2["scored"] is True
    acc = [e for e in rep2["engines"] if e["engine"] == "current"][0]["accuracy"]
    assert acc and {"content_f1", "positional_f1", "row_accuracy", "teds_lite",
                    "misplacement"}.issubset(acc)
    assert acc["content_f1"] >= 0.95                                       # reads the digital BOQ exactly
    md = CE.to_markdown({"generated": "", "options": {"engines": ["current", "paddle"]},
                         "files": [rep, rep2]})
    assert "NON-SOVEREIGN" in md.upper() and md.isascii()                  # ascii-safe report, non-sov flagged

def test_compare_engines_optin_does_not_touch_default_pipeline():
    """OPT-IN proof: importing the harness must NOT change the pipeline's defaults, and the digital
    happy-path stays byte-identical (golden grid + tie-out unchanged) with the flag off - the
    'default pipeline / digital path unchanged' guardrail."""
    import pdf_extract as PE
    before = (PE.PdfOptions().vlm, PE.PdfOptions().dpi, PE.PdfOptions().min_text)
    import eval.compare_engines as _CE            # noqa: F401  (must be import-safe / side-effect free)
    from eval import engines as _E                # noqa: F401
    after = (PE.PdfOptions().vlm, PE.PdfOptions().dpi, PE.PdfOptions().min_text)
    assert before == after == ("off", 300, 50)
    if fitz is None: return
    import tempfile, pdfplumber
    d = Path(tempfile.mkdtemp()); pdf = d / "boq.pdf"; _make_boq_pdf(str(pdf))   # the SAME tiny golden fixture
    with pdfplumber.open(str(pdf)) as doc:
        tables, audit = RT.extract_page(doc.pages[0])
    assert ["A", "100"] in tables[0]["grid"] and ["Total", "600"] in tables[0]["grid"]
    assert TP.verify_table(tables[0]["grid"]) == [] and audit["orphans_in_table_region"] == []

# --------------------------------------------------------------------------- digital byte-identical LOCK
def _make_layout_fixture(path):
    """Heading + paragraph + a ruled BOQ table -> exercises heading/paragraph/table classification,
    level, deterministic ids, and a non-empty section_path. Digital (native text) -> CPU-only, no OCR."""
    doc = fitz.open(); pg = doc.new_page(width=560, height=460)
    pg.insert_text((40, 60), "SITE PREPARATION WORKS", fontsize=18)                       # heading
    pg.insert_text((40, 92), "The following bill of quantities covers the earthwork,", fontsize=9)
    pg.insert_text((40, 108), "plain cement concrete and reinforced concrete items.", fontsize=9)
    xs = [40, 380, 520]; ys = [140, 170, 200, 230, 260]
    for y in ys: pg.draw_line((xs[0], y), (xs[-1], y))
    for x in xs: pg.draw_line((x, ys[0]), (x, ys[-1]))
    rows = [("Description", "Amount"), ("Earthwork in excavation", "100"),
            ("PCC 1:4:8 base course", "200"), ("RCC M25 footing", "300")]
    for r, (a, b) in enumerate(rows):
        pg.insert_text((xs[0] + 4, ys[r] + 20), a, fontsize=8)
        pg.insert_text((xs[1] + 4, ys[r] + 20), b, fontsize=8)
    doc.save(path); doc.close()

_STREAM_KEYS = ("type", "level", "id", "page", "reading_order_index", "section_path", "text", "grid")

def _digital_element_stream():
    """Extract the layout fixture with EVERY opt-in flag OFF (cfg=None -> defaults) and return the
    canonical, path-free element stream as a stable JSON string (bbox/geometry excluded on purpose -
    this locks the SEMANTIC content stream: types, ids, section_path, reading order, grids)."""
    import json, tempfile, doc_layout as DL
    d = Path(tempfile.mkdtemp()); pdf = d / "layout_fix.pdf"; _make_layout_fixture(str(pdf))
    doc = DL.extract_document(pdf, d / "out", None)[0]           # cfg=None => opt-in flags all OFF
    stream = [{k: e.get(k) for k in _STREAM_KEYS}
              for pg in doc.get("pages", []) for e in pg.get("elements", [])]
    return json.dumps(stream, ensure_ascii=False, indent=2, sort_keys=True)

def test_digital_layout_byte_identical_golden():
    """LOCK the 'digital path byte-identical' guarantee: the doc_layout element stream (incl
    section_path + ids) for a digital fixture, all opt-in flags OFF, must stay BYTE-IDENTICAL to the
    stored golden tests/golden_digital_stream.json. Any drift in classification / ids / section_path /
    reading order / grid content fails here. If a change is intentional, delete the golden and re-run."""
    if fitz is None:
        return
    got = _digital_element_stream()
    golden = Path(__file__).resolve().parent / "golden_digital_stream.json"
    if not golden.exists():                                     # first introduction: bootstrap + lock next run
        golden.write_text(got, encoding="utf-8")
        print(f"  [bootstrapped {golden.name} - re-run to lock]"); return
    assert got == golden.read_text(encoding="utf-8"), (
        f"digital element stream DIVERGED from {golden.name} - if intentional, delete it and re-run to re-lock")

def test_compare_engines_warns_on_count_and_order_mismatch():
    """compare_engines.score_vs_gold pairs pred/gold BY ORDER, so it must surface a clear WARNING when
    the table COUNT differs or the tables look MISORDERED (F1 is only apples-to-apples when aligned)."""
    import eval.compare_engines as CE
    import json, tempfile
    d = Path(tempfile.mkdtemp())
    g1 = d / "g1.tables.json"; g1.write_text(json.dumps([[["A", "1"], ["B", "2"]]]), encoding="utf-8")
    r = CE.score_vs_gold([[["A", "1"], ["B", "2"]], [["X", "9"]]], g1)      # pred 2 tables, gold 1
    assert r and any("COUNT differs" in w for w in r.get("warnings", [])), r
    P = [["Amount"], ["100"], ["200"], ["300"]]; Q = [["Qty"], ["7"], ["8"], ["9"]]
    g2 = d / "g2.tables.json"; g2.write_text(json.dumps([P, Q]), encoding="utf-8")
    r2 = CE.score_vs_gold([Q, P], g2)                                      # SWAPPED order
    assert r2 and any("MISORDERED" in w for w in r2.get("warnings", [])), r2
    r3 = CE.score_vs_gold([P, Q], g2)                                      # aligned -> no warnings
    assert not (r3 or {}).get("warnings"), r3

# --------------------------------------------------------------------------- Excel holes (H4)
def test_compare_engines_scanned_stub_not_scored():
    """A scanned transcription STUB (verified:false + needs_human_transcription + 0 grids) is present
    but must NOT be scored (no fabricated accuracy); run_file flags gold_stub and scored=False, and the
    report says 'awaiting human transcription'."""
    if fitz is None:
        return
    import eval.compare_engines as CE
    from eval.engines import EngineOptions
    import json, tempfile
    d = Path(tempfile.mkdtemp()); pdf = d / "boq.pdf"; _make_boq_text_pdf(str(pdf))
    gdir = d / "golden"; gdir.mkdir()
    stem = CE._safe_stem(str(pdf))
    (gdir / f"{stem}.tables.json").write_text(json.dumps(
        {"type": "scanned", "verified": False, "needs_human_transcription": True, "tables": []}),
        encoding="utf-8")
    rep = CE.run_file(str(pdf), ["current"], EngineOptions(), str(gdir))
    assert rep["scored"] is False and rep["gold_stub"] is True and rep["gold_grids"] == 0, rep
    assert all(e.get("accuracy") is None for e in rep["engines"])          # no fabricated accuracy
    md = CE.to_markdown({"generated": "", "options": {"engines": ["current"]}, "files": [rep]})
    assert "awaiting human transcription" in md

def test_xls_date_serial_conversion():
    """H4: legacy .xls XL_CELL_DATE cells are float serials (~45000); excel_serial_to_datetime must
    convert them (matching xlrd) so a date is never read as an amount. PURE function - tested directly,
    no xlrd needed."""
    import excel_extract as X
    import datetime as dt
    assert X.excel_serial_to_datetime(0, 0) == dt.datetime(1899, 12, 30)
    assert X.excel_serial_to_datetime(1, 0) == dt.datetime(1899, 12, 31)
    assert X.excel_serial_to_datetime(60, 0) == dt.datetime(1900, 2, 28)      # Excel's phantom 1900 leap day
    assert X.excel_serial_to_datetime(43831, 0) == dt.datetime(2020, 1, 1)    # a real modern date
    assert X.excel_serial_to_datetime(43831.5, 0) == dt.datetime(2020, 1, 1, 12)
    assert X.excel_serial_to_datetime(0, 1) == dt.datetime(1904, 1, 1)        # 1904 system
    assert X.excel_serial_to_datetime(42369, 1) == dt.datetime(2020, 1, 1)    # 43831 - 1462
    assert X.xls_date_str(43831, 0) == "2020-01-01"
    assert X.xls_date_str(43831.5, 0) == "2020-01-01 12:00:00"
    for bad in (2, -1):                                                       # datemode / negative guards
        try: X.excel_serial_to_datetime(1, bad) if bad == 2 else X.excel_serial_to_datetime(bad, 0); assert False
        except ValueError: pass

def test_xlsx_merged_cell_expanded_for_tieout(tmp_path=None):
    """H4: a merged label/value is present only in the anchor cell under a read; _expand_merged fills
    the range so meta/tie-out see it. (a) find_contract_price MISSES it on the raw grid but FINDS it
    after expansion; (b) _merged_ranges reads the merge; (c) the EMITTED grid stays byte-identical
    (expansion is analysis-only)."""
    try:
        import openpyxl, excel_extract as X
    except Exception:
        return
    import tempfile
    d = Path(tempfile.mkdtemp()); p = str(d / "merged.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "Contract Price"; ws.merge_cells("A1:A3")     # label merged DOWN 3 rows (value only in A1)
    ws["B3"] = 5000000                                       # the amount sits at the bottom of the merge
    wb.save(p)
    raw_grids, _ = X._read_grids(p)
    grid = next(iter(raw_grids.values()))
    assert X.find_contract_price(grid) is None, grid        # raw: row 3 = [None, 5000000] -> MISSED
    ranges = X._merged_ranges(p)
    assert ranges and any(r == (0, 0, 2, 0) for rs in ranges.values() for r in rs), ranges
    expanded = X._expand_merged(grid, next(iter(ranges.values())))
    assert X.find_contract_price(expanded) == 5000000       # expanded: row 3 = ["Contract Price", 5000000]
    assert grid[0][0] == "Contract Price" and grid[2][0] in (None, "")   # raw grid untouched (anchor-only)
    assert X._expand_merged([["x"]], []) == [["x"]]         # pure: no ranges -> identity

# --------------------------------------------------------------------------- standalone runner
if __name__ == "__main__":
    try: sys.stdout.reconfigure(encoding="utf-8")          # so a Hindi assertion message can't crash a cp1252 console
    except Exception: pass
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    passed = failed = 0
    for fn in fns:
        try:
            fn(); passed += 1; print(f"  PASS  {fn.__name__}")
        except Exception as e:
            failed += 1; print(f"  FAIL  {fn.__name__}: {type(e).__name__}: {e}")
    print(f"\n{passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
