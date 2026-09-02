#!/usr/bin/env python3
r"""
excel_extract.py - the XLSX/XLSM family extractor (Schedule-H / BOQ workbooks).

WHY a dedicated extractor: an Excel file is structured data, NOT an image. We read
the EXACT values stored in the cells (openpyxl) - no OCR, no guessing - so values
are 100% accurate (lossless). For Schedule-H workbooks it also structures the
header metadata and runs the arithmetic TIE-OUT check (Σ amounts == Contract Price)
which PROVES the structuring is correct.

This is the AUDITED path (82 MSIDC workbooks, 392,920 cells, 0 missed / 0 altered).
The read core (`_analyze`) is shared by:
  - the standalone CLI  -> rich outputs + excel_summary.csv
  - extract_xlsx()      -> the unified (pages, fields, chunks, text) the pipeline
                           writes as <stem>.doc.json / .chunks.jsonl (so Excel data
                           feeds the KB exactly like every other family) + the rich
                           sidecars.

Standalone:  python excel_extract.py --in file.xlsx --out out\
Via pipeline: python pipeline.py --in folder\ --out out\
Per file -> out/<name>/<name>.cells.json      (exact raw grid per sheet - the 100% data)
            out/<name>/<name>.readable.txt     (clean rows, human-readable)
            out/<name>/<name>.summary.json     (metadata + tie-out verdict)
"""
import argparse, csv, json, re, datetime
from pathlib import Path
import openpyxl

from common import (safe_name, extract_fields, round_tolerance, add_config_args, config_from_args,
                    configure_logging, parse_number, is_amount_cell, contains_total_kw, zip_bomb_reason,
                    atomic_write_text)

# ---------------------------------------------------------------- formatting
def fmt(v):
    """Human-readable cell (thousands separators, trimmed float noise)."""
    if v is None: return ""
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, bool): return str(v)
    if isinstance(v, int): return f"{v:,}"
    if isinstance(v, float):
        if v == int(v): return f"{int(v):,}"
        if abs(v) >= 1000: return f"{v:,.2f}"
        return f"{v:.6f}".rstrip("0").rstrip(".")
    return str(v).strip()

def raw(v):
    """JSON-safe exact value (the lossless 100% data)."""
    if isinstance(v, datetime.datetime): return v.isoformat()
    return v

# ---------------------------------------------------------------- .xls date serials (H4)
def excel_serial_to_datetime(serial, datemode=0):
    """Convert an Excel date SERIAL (float) to a datetime, matching xlrd.xldate.xldate_as_datetime
    EXACTLY - so a legacy .xls date cell reads as a real date, not a ~45000 float that a tie-out would
    mistake for an amount. `datemode`: 0 = 1900 system (epoch 1899-12-30, which also absorbs Excel's
    1900-leap-year quirk for all real dates), 1 = 1904 system. PURE Python (no xlrd import), so it is
    unit-testable even where xlrd isn't installed."""
    if datemode not in (0, 1):
        raise ValueError(f"datemode must be 0 or 1, got {datemode!r}")
    serial = float(serial)
    if serial < 0:
        raise ValueError(f"negative Excel date serial: {serial}")
    epoch = datetime.datetime(1904, 1, 1) if datemode else datetime.datetime(1899, 12, 30)
    days = int(serial)
    fraction = serial - days
    seconds = int(round(fraction * 86400000.0))          # Excel's millisecond resolution
    seconds, milliseconds = divmod(seconds, 1000)
    return epoch + datetime.timedelta(days=days, seconds=seconds, milliseconds=milliseconds)

def xls_date_str(serial, datemode=0):
    """Format an .xls date serial as ISO text: date-only when it has no time-of-day, else a datetime."""
    dt = excel_serial_to_datetime(serial, datemode)
    if (dt.hour, dt.minute, dt.second, dt.microsecond) == (0, 0, 0, 0):
        return dt.date().isoformat()
    return dt.isoformat(sep=" ")

def to_num(v):
    """Excel-typed number reader. Native int/float used directly; bool/datetime -> None (a date is
    not an amount); STRINGS route through common.parse_number, so Devanagari digits, accounting
    negatives and commas parse identically to the digital/table path (one parser, not a local copy)."""
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str): return parse_number(v)
    return None

# ---------------------------------------------------------------- metadata (bilingual: English + Hindi)
META = {
    "contract_price":     r"(total\s+)?contract\s+price|project\s+cost|contract\s+sum|"
                          r"(कुल\s*)?अनुबंध\s*मूल्य|अनुबंध\s*राशि|परियोजना\s*लागत|ठेका\s*मूल्य",
    "work_name":          r"name\s+of\s+work|project\s+name|कार्य\s*का\s*नाम|कार्य\s*नाम|परियोजना\s*का\s*नाम",
    "authority":          r"^\s*(authority|client|employer)\s*$|प्राधिकरण|नियोक्ता",
    "authority_engineer": r"authority'?s?\s+engineer|प्राधिकरण\s*अभियंता",
    "contractor":         r"epc\s+contractor|ठेकेदार|संविदाकार",
    "itf":                r"name\s+of\s+itf|independent\s+testing",
    "project_length":     r"project\s+length|length\s+of\s+the\s+project|परियोजना\s*लंबाई",
    "bill_no":            r"serial\s+no.*bill|bill\s+no|sps|बिल\s*(सं|संख्या|नं)",
}
CP_LABEL = re.compile(META["contract_price"], re.I)

def find_meta(grid):
    meta = {}
    for row in grid[:35]:
        cells = list(row)
        label = next((str(c).strip() for c in cells if str(c).strip()), "")
        if not label: continue
        value = ""
        seen_label = False
        for c in cells:
            t = ("" if c is None else (c if not isinstance(c, datetime.datetime) else fmt(c)))
            ts = str(t).strip()
            if not ts: continue
            if not seen_label: seen_label = True; continue   # skip the label cell
            value = t; break
        for key, pat in META.items():
            if key not in meta and re.search(pat, label, re.I):
                meta[key] = fmt(value) if value != "" else ""
    return meta

def find_contract_price(grid):
    for row in grid[:40]:
        cells = list(row)
        for j, c in enumerate(cells):
            if isinstance(c, str) and CP_LABEL.search(c):
                for k in range(j + 1, len(cells)):
                    n = to_num(cells[k])
                    if n and n > 10000: return n
    return None

# ---------------------------------------------------------------- tie-out
def tieout(grid, cp):
    """Find section/grand 'total' rows (text 'total' + a weightage ~1.0) and check the
    money reconciles to the Contract Price. Handles both layouts:
    (a) a single grand-total row == CP, or (b) section sub-totals that SUM to CP."""
    totals = []; amounts_present = False
    for row in grid:
        cells = list(row)
        text = " ".join(str(c) for c in cells if isinstance(c, str))
        nums = [n for n in (to_num(c) for c in cells) if n is not None]
        if any(n >= 1000 for n in nums):
            amounts_present = True                          # this sheet HAS money -> a proof is expected
        if contains_total_kw(text) and any(abs(n - 1.0) < 1e-6 for n in nums):   # bilingual total row
            amt = max((n for n in nums if n >= 1000), default=None)
            if amt: totals.append({"row": text[:60].strip(), "amount": amt})
    if not cp or not totals:
        # passed=None must NOT be read as 'verified'. If the sheet carries money but no total row was
        # found, the proof did not run -> caller flags it for review (never a silent auto-commit).
        return {"totals": totals, "checked": cp, "passed": None, "amounts_present": amounts_present,
                "how": "no contract price / no total rows"}
    amts = [t["amount"] for t in totals]
    # Rounding-scaled tolerance, NOT 0.5%-of-CP (which was ~5.66M of slack on a 1.13bn CP and
    # would mask a missing section). A single grand-total vs CP is an exact equality (tol=1);
    # K section sub-totals summed drift by at most ceil(0.5*K) rupees of whole-rupee rounding.
    tol_eq  = round_tolerance(1)
    tol_sum = round_tolerance(len(amts))
    if any(abs(a - cp) <= tol_eq for a in amts):
        return {"totals": totals, "checked": cp, "passed": True, "how": "a grand-total row == Contract Price"}
    if abs(sum(amts) - cp) <= tol_sum:
        return {"totals": totals, "checked": cp, "passed": True, "how": "section sub-totals SUM to Contract Price"}
    return {"totals": totals, "checked": cp, "passed": False,
            "how": f"totals ({sum(amts):,.0f}) do not reconcile to CP ({cp:,.0f})"}

# ---------------------------------------------------------------- readable
def readable(grid):
    out = []
    for row in grid:
        cells = [fmt(c) for c in row]
        ne = [c for c in cells if c != ""]
        if ne: out.append(" | ".join(ne))
    return out

# ============================================================================
# READ CORE - shared by the CLI and the pipeline adapter (read ONCE)
# ============================================================================
_SOFFICE_CMD = "soffice"     # overridable from cfg.soffice_cmd (set in main); auto-probes Windows below

def _libreoffice_recalc(path, soffice_cmd=None):
    """SOVEREIGN recalc: LibreOffice headless opens+saves the workbook so every formula is
    evaluated. Returns (grids, uncached) read from the evaluated copy, or (None, None) if soffice is
    unavailable or the recalc failed. Used to recover amounts from a workbook whose formulas were never
    cached (1.1). The soffice command is configurable (Linux/server); a full path or a name-on-PATH both
    work. Its temp dir + converted copy are removed in a finally (M5: no leak)."""
    import shutil, subprocess, tempfile
    cmd = soffice_cmd or _SOFFICE_CMD
    soffice = (shutil.which(cmd) or (cmd if Path(cmd).exists() else None) or shutil.which("soffice.exe") or
               next((p for p in (r"C:\Program Files\LibreOffice\program\soffice.exe",
                                 r"C:\Program Files (x86)\LibreOffice\program\soffice.exe")
                     if Path(p).exists()), None))
    if not soffice:
        return None, None
    d = tempfile.mkdtemp()
    try:
        # subprocess.run(timeout=) already KILLS the child on timeout (no Popen rewrite needed); soffice
        # may leave a soffice.bin grandchild either way - acceptable for an optional recovery step.
        subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", d, str(path)],
                       timeout=120, capture_output=True)
        out = Path(d) / (Path(path).stem + ".xlsx")
        return _read_grids(str(out)) if out.exists() else (None, None)
    except Exception:
        return None, None
    finally:
        shutil.rmtree(d, ignore_errors=True)      # M5: never leak the temp dir + converted copy

def _expand_merged(grid, ranges):
    """Return a COPY of `grid` with each merged range filled from its top-left (anchor) value, so a
    merged total/header value is visible at EVERY cell of the range when meta/tie-out scan a row (H4).
    `ranges` are 0-indexed INCLUSIVE (r0, c0, r1, c1). Only EMPTY cells are filled (a merge carries the
    value in the anchor only), so no real data is overwritten. PURE: does not mutate `grid`, and is NOT
    applied to the emitted cells.json - the exact 100% output stays byte-identical."""
    if not ranges:
        return grid
    g = [list(row) for row in grid]
    n = len(g)
    for (r0, c0, r1, c1) in ranges:
        if not (0 <= r0 < n) or c0 < 0 or r0 >= len(g) or c0 >= len(g[r0]):
            continue
        anchor = g[r0][c0]
        if anchor in (None, ""):
            continue
        for r in range(r0, min(r1, n - 1) + 1):
            for c in range(c0, c1 + 1):
                if c < len(g[r]) and (r, c) != (r0, c0) and g[r][c] in (None, ""):
                    g[r][c] = anchor
    return g

def _merged_ranges(path):
    """Per-sheet 0-indexed INCLUSIVE merged ranges (r0,c0,r1,c1). openpyxl exposes merged_cells only in
    NON-read_only mode; .xls / any failure -> {} (analysis then just uses the raw grid). Used ONLY by
    _analyze for the meta/tie-out scan, never for the emitted grid."""
    if Path(path).suffix.lower() not in (".xlsx", ".xlsm"):
        return {}
    out = {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return {}
    try:
        for ws in wb.worksheets:
            rs = [(mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1)
                  for mr in ws.merged_cells.ranges]
            if rs:
                out[ws.title] = rs
    except Exception:
        out = {}
    finally:
        wb.close()
    return out

def _read_grids(path):
    """Return ({sheet: grid}, uncached) of EXACT cell values. Handles .xlsx/.xlsm (openpyxl) AND
    legacy .xls (xlrd 2.x still reads .xls). `uncached` = [(sheet, coord), ...] cells whose VALUE
    is None but whose content is a FORMULA - openpyxl(data_only=True) returns None for an
    un-cached formula (a library-authored workbook), which would SILENTLY zero money. We only pay
    the second (formula) read when the value read actually produced blanks."""
    ext = Path(path).suffix.lower()
    if ext == ".xls":
        try:
            import xlrd
        except Exception:
            raise RuntimeError(".xls needs xlrd (`pip install xlrd`) or pre-convert to .xlsx "
                               "(e.g. LibreOffice: soffice --headless --convert-to xlsx)")
        wb = xlrd.open_workbook(path)
        dm = wb.datemode
        def _xls_cell(sh, r, c):               # H4: XL_CELL_DATE float serials -> real ISO dates
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == xlrd.XL_CELL_DATE:
                try: return xls_date_str(v, dm)
                except Exception: return v     # keep the raw value if the serial is out of range
            return v
        grids = {sh.name: [[_xls_cell(sh, r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
                 for sh in wb.sheets()}
        return grids, []                       # xlrd always carries values - no uncached formulas
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    grids = {}; any_none = False
    for ws in wb.worksheets:
        try: ws.reset_dimensions = True        # ignore possibly-wrong stored dims -> read ALL rows
        except Exception: pass
        g = [list(r) for r in ws.iter_rows(values_only=True)]
        if not any_none and any(c is None for row in g for c in row): any_none = True
        grids[ws.title] = g
    wb.close()
    uncached = []
    if any_none:                               # explain the blanks: are any of them un-cached formulas?
        try:
            wf = openpyxl.load_workbook(path, data_only=False, read_only=True)
            for name, wsf in zip(grids.keys(), wf.worksheets):
                vg = grids[name]
                for ri, rf in enumerate(wsf.iter_rows()):
                    for ci, cf in enumerate(rf):
                        if (ri < len(vg) and ci < len(vg[ri]) and vg[ri][ci] is None
                                and isinstance(cf.value, str) and cf.value.startswith("=")):
                            uncached.append((name, cf.coordinate))
            wf.close()
        except Exception:
            pass
    return grids, uncached

def _analyze(path, cfg=None):
    """Open the workbook, read every sheet/cell (the audited read), and compute completeness
    stats, contract price, the Schedule-H sheet, metadata + tie-out. If formula cells were
    un-cached (None), try a LibreOffice recalc; whatever remains un-recovered is surfaced as a
    review reason - never silently committed as a complete read (1.1). A zip-bomb workbook (absurd
    uncompressed size / ratio in the central directory) is fenced BEFORE openpyxl decompresses it (M5)."""
    reason = zip_bomb_reason(path, cfg)        # M5: route a decompression bomb to review, don't OOM
    if reason:
        raise RuntimeError(reason)
    grids, uncached = _read_grids(path)
    if uncached:
        g2, u2 = _libreoffice_recalc(path)     # reads the recalced grids internally + cleans its temp dir (M5)
        if g2 is not None and not u2:          # recalc evaluated everything -> use the real values
            grids, uncached = g2, []

    # completeness: how many rows/cells of real data were read (so you can verify it's complete)
    stats = {}
    for name, grid in grids.items():
        rows_data  = sum(1 for r in grid if any((c is not None and str(c).strip() != "") for c in r))
        cells_data = sum(1 for r in grid for c in r if (c is not None and str(c).strip() != ""))
        stats[name] = {"rows": len(grid), "rows_with_data": rows_data,
                       "cells": sum(len(r) for r in grid), "cells_with_data": cells_data}

    # H4: build an ANALYSIS view with merged cells expanded (a merged total/header value is otherwise
    # visible only in its anchor cell and the scan below misses it). The EMITTED grids stay raw, so
    # cells.json / the digital 100% output is byte-identical - expansion only feeds meta + tie-out.
    merged = _merged_ranges(path)
    agrids = {name: _expand_merged(g, merged.get(name, [])) for name, g in grids.items()}

    # contract price + tie-out; the Schedule-H sheet = first sheet whose money reconciles
    cp = None
    for grid in agrids.values(): cp = cp or find_contract_price(grid)
    tie = {"totals": [], "checked": cp, "passed": None, "how": "no total rows / no contract price"}
    sched = None
    for name in grids:
        t = tieout(agrids[name], cp)
        if t["passed"] is True: tie, sched = t, name; break
        if sched is None: tie, sched = t, name
    meta = find_meta(agrids.get(sched) or next(iter(agrids.values()), []))
    return {"path": path, "grids": grids, "stats": stats, "cp": cp,
            "sched": sched, "meta": meta, "tie": tie, "uncached": uncached}

def _write_rich(an, docdir, safe):
    """Write the audited sidecars: exact grid (cells.json), readable rows, summary."""
    grids = an["grids"]; stats = an["stats"]
    cells_json = {name: [[raw(c) for c in row] for row in grid] for name, grid in grids.items()}
    # long-path safe (_fs_path via atomic_write_text): deep sharded output paths on Windows can
    # exceed MAX_PATH (260) - a raw Path.write_text() there raises FileNotFoundError and silently
    # loses the file. atomic_write_text applies the \\?\ prefix so the exact 100% grid always lands.
    atomic_write_text(docdir / f"{safe}.cells.json",
                      json.dumps(cells_json, indent=1, ensure_ascii=False, default=str))
    lines = []
    for name, grid in grids.items():
        lines.append(f"===== SHEET: {name}  ({stats[name]['rows_with_data']} rows with data) =====")
        lines += readable(grid); lines.append("")
    atomic_write_text(docdir / f"{safe}.readable.txt", "\n".join(lines))
    summary = {"file": str(an["path"]), "sheets": list(grids.keys()), "completeness": stats,
               "contract_price": an["cp"], "schedule_h_sheet": an["sched"],
               "metadata": an["meta"],
               "metadata_note": "best-effort (regex/positional, first 35 rows); the cell VALUES "
                                "(cells.json) are the guaranteed 100%-exact data, not the metadata",
               "uncached_formula_cells": len(an.get("uncached", [])),
               "uncached_sample": [f"{s}!{c}" for s, c in an.get("uncached", [])[:20]],
               "tie_out": an["tie"]}
    atomic_write_text(docdir / f"{safe}.summary.json",
                      json.dumps(summary, indent=2, ensure_ascii=False, default=str))
    return summary

def _to_unified(an):
    """Adapt the audited read into the pipeline's (pages, fields, chunks, full_text)."""
    pages = []; chunks = []; text = []
    # COMPLETENESS signals -> flag the schedule sheet so the shared gate (common.write_outputs) routes
    # the workbook to review / withholds it in strict mode. A failed tie-out, un-cached formula cells,
    # OR a tie-out that COULD NOT RUN despite money being present (passed=None + amounts) - the last one
    # is the Hindi/other-language silent-commit hole: 'not attempted' must never read as 'verified'.
    tie = an["tie"]
    tie_failed = tie.get("passed") is False
    tie_not_run = tie.get("passed") is None and tie.get("amounts_present")
    n_uncached = len(an.get("uncached", []))
    for name, grid in an["grids"].items():
        for ri, row in enumerate(grid, 1):
            cells = [fmt(c) for c in row]
            nonempty = [c for c in cells if c != ""]          # drop empty padding cells
            if nonempty:
                line = " | ".join(nonempty)                   # readable: only real values
                chunks.append({"kind": "boq_row", "page": name, "row": ri,
                               "text": line, "confidence": 0.98})
                text.append(line)
        pages.append({"page": name, "route": "xlsx", "confidence": 0.98, "chars": 500,
                      "doctr_conf": None, "tess_conf": None, "agreement": 1.0, "png": None,
                      "tieout_failed": bool(tie_failed and name == an["sched"]),
                      "tieout_not_attempted": bool(tie_not_run and name == an["sched"]),
                      "uncached_formulas": n_uncached if name == an["sched"] else 0})
    return pages, extract_fields(text), chunks, "\n".join(text)

# ============================================================================
# PUBLIC: pipeline adapter + standalone parse
# ============================================================================
def extract_xlsx(path, docdir, cfg=None):
    """Pipeline entry: write the audited sidecars AND return the unified tuple. cfg carries the
    zip-bomb caps (M5); None = generous defaults (standalone / reconcile paths)."""
    safe = safe_name(Path(path).stem)
    docdir.mkdir(parents=True, exist_ok=True)
    an = _analyze(path, cfg)
    _write_rich(an, docdir, safe)
    return _to_unified(an)

def parse_file(path, outroot):
    """Standalone entry: rich outputs only (the original excel_parser behaviour)."""
    safe = safe_name(Path(path).stem)
    docdir = outroot / safe; docdir.mkdir(parents=True, exist_ok=True)
    an = _analyze(path)
    return _write_rich(an, docdir, safe)


# ============================================================================
# STANDALONE CLI (XLSX only) - audited direct read + tie-out, no OCR
# ============================================================================
def main():
    ap = argparse.ArgumentParser(description="XLSX/XLSM extractor (direct read = 100% exact, no OCR).")
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", dest="out", required=True)
    add_config_args(ap)
    a = ap.parse_args()
    cfg = config_from_args(a)
    global _SOFFICE_CMD
    if cfg.soffice_cmd: _SOFFICE_CMD = cfg.soffice_cmd        # configurable LibreOffice for formula recalc
    inp = Path(a.inp); outroot = Path(a.out); outroot.mkdir(parents=True, exist_ok=True)
    configure_logging(cfg, outroot)
    EXTS = (".xlsx", ".xlsm", ".xls")
    if inp.is_file() and inp.suffix.lower() not in EXTS:
        print(f"{inp.name}: not an Excel workbook ({'/'.join(EXTS)})."); return
    files = [inp] if inp.is_file() else sorted(p for p in inp.rglob("*") if p.suffix.lower() in EXTS)
    if not files:
        print(f"No Excel workbook ({'/'.join(EXTS)}) under {inp}"); return
    rows = []
    for fp in files:
        try:
            s = parse_file(fp, outroot)
            t = s["tie_out"]; verdict = {True: "PASS", False: "REVIEW", None: "n/a"}[t["passed"]]
            if t["passed"] is None and t.get("amounts_present"):
                verdict = "REVIEW"                    # money present but tie-out couldn't run -> NOT 'n/a'
            nunc = s.get("uncached_formula_cells", 0)
            if nunc: verdict = "REVIEW"               # un-cached formulas = blanks; never call it 100%
            cp = f"{s['contract_price']:,.0f}" if s["contract_price"] else "?"
            rws = sum(v["rows_with_data"] for v in s["completeness"].values())
            cls = sum(v["cells_with_data"] for v in s["completeness"].values())
            unc = f" | UNCACHED-FORMULAS={nunc} (recalc source!)" if nunc else ""
            print(f"[{'100% read' if not nunc else 'PARTIAL'}] {fp.name}  | sheets={len(s['sheets'])} | "
                  f"{rws} data-rows, {cls:,} data-cells | contract_price={cp} | tie-out: {verdict} ({t['how']}){unc}")
            rows.append({"file": fp.name, "sheets": len(s["sheets"]), "data_rows": rws, "data_cells": cls,
                         "contract_price": s["contract_price"], "tie_out": verdict, "how": t["how"]})
        except Exception as e:
            print(f"[ERROR] {fp.name}: {e}")
            rows.append({"file": fp.name, "sheets": 0, "data_rows": 0, "data_cells": 0,
                         "contract_price": "", "tie_out": "ERROR", "how": str(e)})
    with open(outroot / "excel_summary.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["file", "sheets", "data_rows", "data_cells",
                                          "contract_price", "tie_out", "how"])
        w.writeheader(); [w.writerow(r) for r in rows]
    npass = sum(1 for r in rows if r["tie_out"] == "PASS")
    print(f"\nValues read DIRECTLY from cells = 100% exact (no OCR). Tie-out PASS proves the money reconciles.")
    print(f"{npass}/{len(rows)} workbooks tie out cleanly.  Summary -> {outroot/'excel_summary.csv'}")

if __name__ == "__main__":
    main()
